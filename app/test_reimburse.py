# -*- coding: utf-8 -*-
"""报销汇总模块测试：直接 python3 运行，无 pytest 依赖。"""
import os
import shutil
import sys
import tempfile

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import reimburse
import reimburse_ocr


PASS = 0
FAIL = 0


OIL_TEXT = """
             成品油
                        电子发票（普通发票）                                    发票号码：26327000001152328150
                                                                      开票日期：2026年06月21日
      项目名称      规格型号   单   位          数 量         单   价     金 额        税率/征收率             税     额
*汽油*92号车用汽油（VIB         升       50.82142857            7    355.75       13%                  46.25
）
        合     计                                            ¥ 355.75                        ¥ 46.25
      价税合计(大写)         肆佰零贰圆整                               (小写)       ¥ 402.00
"""


TOLL_TEXT = """
                    江苏省车辆通行费通用（电子）发票                                 发票代码: 132012300111
                                                                     发票号码: 02273714
                                                                     开票日期: 2026 年 06 月 21 日
        货物或应税劳务、服务名称                      数量                            金额（元）
             车辆通行费                        1                               20.00
           合         计                                                    20.00
        价税合计（大写）              贰拾圆                              （小写）       20.00
"""


NEW_TOLL_TEXT = """
                             电子发票（普通发票）                                发票号码： 26337000000587975272
                                                                       开票日期： 2026年06月30日
    项目名称         规格型号      单 位          数 量        单 价         金 额      税率/征收率               税 额
*生产生活服务*代收高      无          无             1        41.00       41.00      不征税
速通行费
       合     计                                                ¥41.00                        ¥0.00
    价税合计（大写）            肆拾壹元整                                 (小写) ¥41.00
"""


PYPDF_OIL_TEXT = """
成品油下载次数：1
电子发票（普通发票） 发票号码：
开票日期：26327000001152328150
2026年06月21日
项目名称 规格型号 单  位 数  量 单  价 金  额 税率/征收率 税  额
*汽油*92号车用汽油（VIB
） 升 50.82142857 7 355.75 13% 46.25
合 计 355.75 46.25 ¥ ¥
价税合计(大写) (小写) 肆佰零贰圆整 ¥402.00
"""


PYPDF_OLD_TOLL_TEXT = """
江苏省车辆通行费通用（电子）发票 发票代码:
发票号码:132012300111
02273714
2026 开票日期:
年月日0621
货物或应税劳务、服务名称  数量 金额（元）
车辆通行费 1 20.00
合          计
价税合计（大写）20.00
贰拾圆 （小写） 20.00
"""


PYPDF_NEW_TOLL_TEXT = """
 发票号码： 26337000000587975272
 开票日期： 2026年06月30日
 开票人: 杨颖电子发票（普通发票）
项目名称 规格型号 单 位 数 量 单 价 金 额税率/征收率 税 额
*生产生活服务*代收高
速通行费无 无 1 41.00 41.00 不征税
合 计 ¥41.00 ¥0.00
价税合计（大写）
注 (小写)¥41.00 肆拾壹元整
"""


def check(name, cond):
    global PASS, FAIL
    if cond:
        PASS += 1
        print('  PASS %s' % name)
    else:
        FAIL += 1
        print('  FAIL %s' % name)


def test_parse_text():
    print('\n[发票文本解析]')
    oil = reimburse._parse_invoice_text(OIL_TEXT, '油费发票/6.21-6.23油费发票.pdf')
    check('油费发票号码', oil.invoice_number == '26327000001152328150')
    check('油费日期', oil.issue_date == '2026-06-21')
    check('油费项目', oil.item_name == '*汽油*92号车用汽油（VIB）')
    check('油费金额税额合计', (oil.amount, oil.tax, oil.total) == (355.75, 46.25, 402.0))
    check('油费状态', oil.status == '已解析')

    toll = reimburse._parse_invoice_text(TOLL_TEXT, '通行费发票/6.21通行费发票1.pdf')
    check('通行费代码', toll.invoice_code == '132012300111')
    check('通行费号码', toll.invoice_number == '02273714')
    check('通行费税额推导', toll.tax == 0.0)
    check('通行费类别', toll.category == 'ETC通行费')

    nt = reimburse._parse_invoice_text(NEW_TOLL_TEXT, '通行费发票/6.30通行费发票1.pdf')
    check('新版通行费无代码允许', nt.invoice_code == '')
    check('新版通行费项目拼接', nt.item_name == '*生产生活服务*代收高速通行费')
    check('新版通行费状态', nt.status == '已解析')

    po = reimburse._parse_invoice_text(PYPDF_OIL_TEXT, '油费发票/pypdf.pdf')
    check('pypdf 油费号码错位修正', po.invoice_number == '26327000001152328150')
    check('pypdf 油费日期合法', po.issue_date == '2026-06-21')
    check('pypdf 油费项目不带合计', po.item_name == '*汽油*92号车用汽油（VIB）')

    pot = reimburse._parse_invoice_text(PYPDF_OLD_TOLL_TEXT, '通行费发票/pypdf-old.pdf')
    check('pypdf 老通行费代码/号码', pot.invoice_code == '132012300111' and pot.invoice_number == '02273714')
    check('pypdf 老通行费日期', pot.issue_date == '2026-06-21')

    pnt = reimburse._parse_invoice_text(PYPDF_NEW_TOLL_TEXT, '通行费发票/pypdf-new.pdf')
    check('pypdf 新通行费价税合计', pnt.total == 41.0)
    check('pypdf 新通行费项目', pnt.item_name == '*生产生活服务*代收高速通行费')


def test_duplicate_and_export():
    print('\n[重复检查与导出]')
    tmp = tempfile.mkdtemp(prefix='reim_test_')
    try:
        r1 = reimburse._parse_invoice_text(TOLL_TEXT, os.path.join(tmp, 'a.pdf'))
        r2 = reimburse._parse_invoice_text(TOLL_TEXT, os.path.join(tmp, 'b.pdf'))
        oil = reimburse._parse_invoice_text(OIL_TEXT, os.path.join(tmp, 'oil.pdf'))
        trips = [reimburse.TripRecord(
            date='2026-06-21',
            start_file=os.path.join(tmp, '6.21出发.jpg'),
            end_file=os.path.join(tmp, '6.21结束.jpg'),
            address='中国江苏省南京市浦口区天华北路1号',
            start_odometer=131066,
            end_odometer=131188,
        )]
        out = reimburse.build_reimburse_outputs([r1, r2, oil], trips, os.path.join(tmp, '输出'), template_dir=tmp)
        check('导出三表', all(os.path.exists(p) for p in out.values()))

        wb = openpyxl.load_workbook(out['invoice'], data_only=False)
        ws = wb.active
        check('电子发票从第4行写入', ws['B4'].value == '江苏省车辆通行费通用（电子）发票')
        check('重复状态标记', '重复' in (ws['K4'].value or '') and '重复' in (ws['K5'].value or ''))
        check('价税合计公式', ws['I4'].value == '=G4+H4')

        wb2 = openpyxl.load_workbook(out['trip'], data_only=False)
        ws2 = wb2.active
        check('行程日期', ws2['B4'].value == '2026-06-21')
        check('行程里程公式', ws2['I4'].value == '=H4-G4')
        check('行程通行费关联', ws2['L4'].value == 40.0)

        wb3 = openpyxl.load_workbook(out['expense'], data_only=True)
        cats = [wb3.active.cell(r, 6).value for r in range(3, 8)]
        check('费用明细含油费', '油费' in cats)
        check('费用明细含通行费', 'ETC通行费' in cats)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_ocr_prefill_parsers():
    print('\n[OCR 预填字段提取]')
    trip = reimburse_ocr.OcrResult(words=[
        reimburse_ocr.OcrWord('2026年6月21日星期日08:37', 1.0, (((44, 1190), (665, 1190), (665, 1242), (44, 1242)))),
        reimburse_ocr.OcrWord('131066km', 0.96, (((480, 377), (596, 384), (593, 424), (476, 417)))),
        reimburse_ocr.OcrWord('中国江苏省南京市浦口区天华北路1号', 0.99, (((71, 2209), (826, 2209), (826, 2261), (71, 2261)))),
    ], text='')
    sug = reimburse_ocr.trip_suggestion(trip)
    check('OCR 行程公里数', sug.odometer == 131066)
    check('OCR 行程地址', sug.address == '中国江苏省南京市浦口区天华北路1号')

    lodging_text = '\n'.join([
        '电子发',
        '发票号码：26332000005313249961',
        '开票日期：2026年06月22日',
        '*生产生活服务*住宿费',
        '181.81',
        '6%',
        '10.91',
        '合',
        '计',
        '￥181.81',
        '￥10.91',
        '价税合计（大写）',
        '(小写)￥192.72',
    ])
    rec = reimburse_ocr.invoice_suggestion(
        reimburse_ocr.OcrResult(words=[reimburse_ocr.OcrWord(x, 0.99) for x in lodging_text.splitlines()], text=lodging_text),
        '住宿费发票/6.21.jpg',
    )
    check('OCR 图片发票类型兜底', rec.invoice_type == '电子发票（普通发票）')
    check('OCR 图片发票号码', rec.invoice_number == '26332000005313249961')
    check('OCR 图片发票金额税额合计', (rec.amount, rec.tax, rec.total) == (181.81, 10.91, 192.72))
    check('OCR 图片发票状态待确认', rec.status == 'OCR预填')


def test_sample_dir_if_present():
    print('\n[真实样例目录扫描]')
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'docs', '报销功能', '报销(1)', '报销'))
    if not os.path.isdir(root):
        print('  SKIP 未发现 docs/报销功能 样例目录')
        return
    report = reimburse.scan_reimburse_dir(root)
    check('样例 PDF 数量', report.pdf_count >= 18)
    check('样例外访行程数量', len(report.trips) >= 8)
    ok_pdf = sum(1 for r in report.invoices if r.source_kind == 'pdf' and r.status in ('已解析', '重复'))
    check('样例 PDF 可解析', ok_pdf >= 18)


def main():
    test_parse_text()
    test_duplicate_and_export()
    test_ocr_prefill_parsers()
    test_sample_dir_if_present()
    print('\nRESULT: %d passed, %d failed' % (PASS, FAIL))
    if FAIL:
        return 1
    print('REIMBURSE TESTS PASSED')
    return 0


if __name__ == '__main__':
    sys.exit(main())
