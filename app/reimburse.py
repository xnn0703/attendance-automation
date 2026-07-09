# -*- coding: utf-8 -*-
"""报销汇总引擎：离线扫描、PDF 发票文本解析、人工复核数据导出。

首版只做确定性文本解析；图片发票和外访截图进入人工复核流程，不做 OCR。
"""
from __future__ import annotations

import datetime as _dt
import os
import re
import shutil
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


IMAGE_EXTS = {'.jpg', '.jpeg', '.png', '.bmp', '.webp'}
PDF_EXTS = {'.pdf'}
OUT_DIR_NAME = '输出'


@dataclass
class InvoiceRecord:
    source_file: str
    invoice_type: str = ''
    issue_date: str = ''
    invoice_code: str = ''
    invoice_number: str = ''
    item_name: str = ''
    amount: Optional[float] = None
    tax: Optional[float] = None
    total: Optional[float] = None
    category: str = ''
    status: str = '待录入'
    error: str = ''
    duplicate_key: str = ''
    source_kind: str = 'pdf'


@dataclass
class TripRecord:
    date: str
    start_file: str = ''
    end_file: str = ''
    address: str = ''
    start_odometer: Optional[int] = None
    end_odometer: Optional[int] = None
    trip_km: Optional[int] = None
    fees: Dict[str, float] = field(default_factory=dict)
    status: str = '需复核'
    error: str = ''


@dataclass
class PendingImage:
    source_file: str
    kind: str
    date: str = ''
    note: str = ''


@dataclass
class ScanReport:
    root: str
    invoices: List[InvoiceRecord]
    trips: List[TripRecord]
    pending_images: List[PendingImage]

    @property
    def pdf_count(self) -> int:
        return sum(1 for r in self.invoices if r.source_kind == 'pdf')

    @property
    def image_invoice_count(self) -> int:
        return sum(1 for r in self.invoices if r.source_kind == 'image')


def extract_pdf_text(path: str) -> str:
    """提取 PDF 文本层；优先 pypdf，缺失/失败时回退到系统 pdftotext。"""
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(path)
        parts = []
        for page in reader.pages:
            parts.append(page.extract_text() or '')
        text = '\n'.join(parts).strip()
        if text:
            return text
    except Exception:
        pass

    exe = shutil.which('pdftotext')
    if not exe:
        raise RuntimeError('无法提取 PDF 文本：未安装 pypdf，且系统没有 pdftotext')
    cp = subprocess.run([exe, '-layout', path, '-'], check=False, capture_output=True, text=True)
    if cp.returncode != 0:
        raise RuntimeError((cp.stderr or 'pdftotext 解析失败').strip())
    return cp.stdout


def parse_invoice_pdf(path: str) -> InvoiceRecord:
    text = extract_pdf_text(path)
    return _parse_invoice_text(text, path)


def _parse_invoice_text(text: str, source_file: str) -> InvoiceRecord:
    lines = [ln.strip() for ln in text.replace('\u3000', ' ').splitlines()]
    compact = '\n'.join(lines)
    rec = InvoiceRecord(source_file=source_file, source_kind='pdf')
    rec.invoice_type = _extract_invoice_type(compact, lines)
    rec.issue_date = _extract_issue_date(compact)
    rec.invoice_code, rec.invoice_number = _extract_code_number(compact)
    rec.item_name = _extract_item_name(lines)
    rec.amount, rec.tax, rec.total = _extract_amounts(lines)
    rec.category = infer_category(source_file, rec.item_name, compact)
    _validate_invoice(rec)
    return rec


def scan_reimburse_dir(root: str) -> ScanReport:
    base = Path(root).expanduser().resolve()
    if not base.is_dir():
        raise FileNotFoundError('报销目录不存在：%s' % root)

    invoices: List[InvoiceRecord] = []
    pending: List[PendingImage] = []
    trip_parts: Dict[Tuple[int, int], Dict[str, str]] = defaultdict(dict)
    image_invoice_mmdd: List[Tuple[InvoiceRecord, Optional[Tuple[int, int]]]] = []

    for path in sorted(base.rglob('*')):
        if not path.is_file() or path.name.startswith('.~'):
            continue
        if OUT_DIR_NAME in path.parts:
            continue
        ext = path.suffix.lower()
        if ext in PDF_EXTS:
            try:
                invoices.append(parse_invoice_pdf(str(path)))
            except Exception as e:
                invoices.append(InvoiceRecord(
                    source_file=str(path), source_kind='pdf', status='需复核',
                    error='PDF 文本解析失败：%s' % e, category=infer_category(str(path), '', ''),
                ))
            continue
        if ext not in IMAGE_EXTS:
            continue

        mmdd = _extract_month_day(path.name)
        name = path.name
        in_wai = any('外访' in part for part in path.parts)
        if in_wai and '出发' in name and mmdd:
            trip_parts[mmdd]['start_file'] = str(path)
        elif in_wai and '结束' in name and mmdd:
            trip_parts[mmdd]['end_file'] = str(path)
        elif _is_invoice_image(path):
            rec = InvoiceRecord(
                source_file=str(path), source_kind='image', issue_date='',
                item_name='', category=infer_category(str(path), '', ''),
                status='待录入', error='图片发票需手工录入后再导出',
            )
            invoices.append(rec)
            image_invoice_mmdd.append((rec, mmdd))
        else:
            pending.append(PendingImage(
                source_file=str(path), kind='图片凭证', date='',
                note='非 PDF 发票/非出发结束截图，保留为待复核图片',
            ))

    year = _infer_year(invoices) or _dt.date.today().year
    for rec, mmdd in image_invoice_mmdd:
        if mmdd:
            rec.issue_date = '%04d-%02d-%02d' % (year, mmdd[0], mmdd[1])
    for img in pending:
        mmdd = _extract_month_day(Path(img.source_file).name)
        if mmdd:
            img.date = '%04d-%02d-%02d' % (year, mmdd[0], mmdd[1])

    trips = []
    for (mo, day), part in sorted(trip_parts.items()):
        trips.append(TripRecord(
            date='%04d-%02d-%02d' % (year, mo, day),
            start_file=part.get('start_file', ''),
            end_file=part.get('end_file', ''),
            status='需复核',
            error='请确认出发/回司公里数和地址',
        ))

    invoices.sort(key=_invoice_sort_key)
    validate_invoice_records(invoices)
    return ScanReport(str(base), invoices, trips, pending)


def validate_invoice_records(records: Sequence[InvoiceRecord]) -> None:
    for rec in records:
        _validate_invoice(rec)
    groups: Dict[str, List[InvoiceRecord]] = defaultdict(list)
    for rec in records:
        key = duplicate_key(rec)
        rec.duplicate_key = key
        if key:
            groups[key].append(rec)
    for key, dupes in groups.items():
        if len(dupes) <= 1:
            continue
        for rec in dupes:
            rec.status = '重复'
            rec.error = _join_error(rec.error, '疑似重复发票：%s' % key)


def duplicate_key(rec: InvoiceRecord) -> str:
    num = (rec.invoice_number or '').strip()
    if not num:
        return ''
    code = (rec.invoice_code or '').strip()
    return '%s-%s' % (code, num) if code else num


def _invoice_sort_key(rec: InvoiceRecord):
    kind_rank = 1 if rec.source_kind == 'image' else 0
    return (rec.issue_date or '9999-99-99', kind_rank, rec.category or '', os.path.basename(rec.source_file))


def build_reimburse_outputs(
    records: Sequence[InvoiceRecord],
    trips: Sequence[TripRecord],
    out_dir: str,
    template_dir: Optional[str] = None,
) -> Dict[str, str]:
    out = Path(out_dir).expanduser().resolve()
    out.mkdir(parents=True, exist_ok=True)
    records = list(records)
    trips = list(trips)
    validate_invoice_records(records)
    _validate_trips(trips)

    invoice_path = str(out / '报销单电子发票登记表（汇总）.xlsx')
    trip_path = str(out / '外访行程记录表（汇总）.xlsx')
    expense_path = str(out / '费用报销明细（付款）（汇总）.xlsx')

    _write_invoice_workbook(records, invoice_path)
    _write_trip_workbook(trips, records, trip_path, template_dir or str(out))
    _write_expense_workbook(records, expense_path, template_dir or str(out))
    return {'invoice': invoice_path, 'trip': trip_path, 'expense': expense_path}


def infer_category(path: str, item_name: str, text: str = '') -> str:
    s = '%s %s %s' % (path, item_name, text)
    if any(x in s for x in ('汽油', '柴油', '油费', '成品油')):
        return '油费'
    if '通行费' in s or '高速' in s or 'ETC' in s.upper():
        return 'ETC通行费'
    if any(x in s for x in ('住宿', '酒店', '宾馆')):
        return '住宿费'
    if '停车' in s:
        return '停车费'
    return '其他'


def _validate_invoice(rec: InvoiceRecord) -> None:
    errors = []
    if rec.source_kind == 'image' and not any((rec.invoice_number, rec.total, rec.item_name)):
        rec.status = '待录入'
        rec.error = rec.error or '图片发票需手工录入后再导出'
        return
    if not rec.invoice_type:
        errors.append('缺发票类型')
    if not rec.issue_date:
        errors.append('缺开票日期')
    if not rec.invoice_number:
        errors.append('缺发票号码')
    if not rec.item_name:
        errors.append('缺项目名称')
    if rec.amount is None:
        errors.append('缺金额')
    if rec.tax is None:
        errors.append('缺税额')
    if rec.total is None:
        errors.append('缺价税合计')
    if rec.amount is not None and rec.tax is not None and rec.total is not None:
        if abs((rec.amount + rec.tax) - rec.total) > 0.01:
            errors.append('金额+税额与价税合计不一致')
    rec.status = '需复核' if errors else '已解析'
    rec.error = '；'.join(errors)


def _validate_trips(trips: Sequence[TripRecord]) -> None:
    for trip in trips:
        errors = []
        if trip.start_odometer is None:
            errors.append('缺出发公里数')
        if trip.end_odometer is None:
            errors.append('缺回司公里数')
        if not trip.address:
            errors.append('缺目的地/地址')
        if trip.start_odometer is not None and trip.end_odometer is not None:
            if trip.end_odometer < trip.start_odometer:
                errors.append('回司公里数小于出发公里数')
                trip.trip_km = None
            else:
                trip.trip_km = trip.end_odometer - trip.start_odometer
                if trip.trip_km > 1000:
                    errors.append('单日里程超过 1000km，请复核')
        trip.status = '需复核' if errors else '已确认'
        trip.error = '；'.join(errors)


def _extract_invoice_type(text: str, lines: Sequence[str]) -> str:
    if '江苏省车辆通行费通用' in text:
        return '江苏省车辆通行费通用（电子）发票'
    if '成品油' in text and '电子发票' in text:
        return '成品油电子发票（普通发票）'
    m = re.search(r'([\u4e00-\u9fa5（）()]*电子发票[（(][^）)]+[）)])', text)
    if m:
        return m.group(1).replace(' ', '')
    for line in lines:
        if '发票' in line and '号码' not in line and '代码' not in line:
            return re.sub(r'\s+', '', line)
    return ''


def _extract_issue_date(text: str) -> str:
    patterns = [
        r'开票日期\s*[:：][^\d]{0,20}(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日',
        r'(\d{4})\s*开票日期\s*[:：]?\s*年\s*月\s*日\s*(\d{2})(\d{2})',
        r'开票日期\s*[:：]\s*(\d{4})(\d{2})(\d{2})(?!\d)',
        r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日',
        r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})',
    ]
    for pat in patterns:
        for m in re.finditer(pat, text):
            y, mo, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 2000 <= y <= 2100 and 1 <= mo <= 12 and 1 <= day <= 31:
                return '%04d-%02d-%02d' % (y, mo, day)
    return ''


def _extract_code_number(text: str) -> Tuple[str, str]:
    code = _first_digits_after(text, '发票代码', min_len=10, max_len=12)
    number = _first_digits_after(text, '发票号码', min_len=8, max_len=24)
    nums = _digit_sequences_near(text, '发票号码', limit=240)
    if nums:
        if code and nums[0] == code and len(nums) >= 2:
            number = nums[1]
        elif len(nums[0]) >= 20:
            number = nums[0]
        elif len(nums) >= 2 and len(nums[0]) >= 10 and 6 <= len(nums[1]) <= 12:
            code = code or nums[0]
            number = nums[1]
        elif not number:
            number = nums[0]
    return code, number


def _first_digits_after(text: str, label: str, min_len: int, max_len: int) -> str:
    for num in _digit_sequences_near(text, label, limit=220):
        if min_len <= len(num) <= max_len:
            return num
    return ''


def _digit_sequences_near(text: str, label: str, limit: int = 180) -> List[str]:
    idx = text.find(label)
    if idx < 0:
        return []
    chunk = text[idx:idx + limit]
    return re.findall(r'\d{%d,%d}' % (6, 24), chunk)


def _extract_item_name(lines: Sequence[str]) -> str:
    for i, line in enumerate(lines):
        s = line.strip()
        if not s.startswith('*'):
            continue
        item = re.split(r'\s{2,}', s, maxsplit=1)[0]
        for j in range(i + 1, min(i + 3, len(lines))):
            nxt = lines[j].strip()
            if not nxt or re.search(r'合\s*计|价税|备注|开票人|税\s*额', nxt):
                break
            prefix = _item_continuation_prefix(nxt)
            if prefix:
                item += prefix
                continue
            if re.search(r'[¥￥]|\d{2,}', nxt) or ':' in nxt or '：' in nxt:
                break
            item += nxt
        return _clean_item(item)
    for line in lines:
        if '车辆通行费' in line:
            return '车辆通行费'
    return ''


def _extract_amounts(lines: Sequence[str]) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    amount = None
    tax = None
    total = None
    for line in lines:
        if '价税合计' in line:
            nums = _money_numbers(line)
            if nums:
                total = nums[-1]
        elif re.search(r'合\s*计', line) and '价税' not in line and '大写' not in line:
            nums = _money_numbers(line)
            if nums:
                amount = nums[0]
                if len(nums) >= 2:
                    tax = nums[1]
    if total is None:
        for line in lines:
            if '小写' in line:
                nums = _money_numbers(line)
                if nums:
                    total = nums[-1]
    if total is not None and amount is None:
        amount = total
    if total is not None and amount is not None and tax is None:
        tax = round(total - amount, 2)
    return amount, tax, total


def _money_numbers(line: str) -> List[float]:
    nums = []
    for m in re.finditer(r'(?:[¥￥]\s*)?([0-9]+(?:\.[0-9]{1,2})?)', line):
        try:
            nums.append(round(float(m.group(1)), 2))
        except ValueError:
            pass
    return nums


def _clean_item(text: str) -> str:
    text = re.sub(r'\s+', '', text)
    return text.replace('（', '(').replace('）', ')').replace('(VIB)', '（VIB）')


def _item_continuation_prefix(text: str) -> str:
    if re.match(r'合\s*计|价税|税\s*额', text):
        return ''
    cut = re.split(r'\s{2,}|无\s+无|\s+\d|[¥￥]|升|个|次|件', text, maxsplit=1)[0].strip()
    if not cut:
        return ''
    if len(cut) <= 12 and re.search(r'[\u4e00-\u9fa5（）()]', cut):
        return cut
    return ''


def _first_match(text: str, pattern: str) -> str:
    m = re.search(pattern, text)
    return m.group(1).strip() if m else ''


def _join_error(a: str, b: str) -> str:
    return ('%s；%s' % (a, b)) if a else b


def _extract_month_day(name: str) -> Optional[Tuple[int, int]]:
    m = re.search(r'(?<!\d)(\d{1,2})[.\-月](\d{1,2})(?!\d)', name)
    if not m:
        return None
    mo, day = int(m.group(1)), int(m.group(2))
    if 1 <= mo <= 12 and 1 <= day <= 31:
        return mo, day
    return None


def _infer_year(records: Sequence[InvoiceRecord]) -> Optional[int]:
    years = []
    for rec in records:
        if rec.issue_date[:4].isdigit():
            years.append(int(rec.issue_date[:4]))
    return Counter(years).most_common(1)[0][0] if years else None


def _is_invoice_image(path: Path) -> bool:
    s = '/'.join(path.parts)
    return '发票' in s


def _find_template(start_dir: str, tokens: Sequence[str], suffixes: Sequence[str] = ('.xlsx',)) -> Optional[str]:
    start = Path(start_dir).expanduser().resolve()
    dirs = [start] + list(start.parents[:8])
    for d in dirs:
        if not d.is_dir():
            continue
        for p in d.iterdir():
            name = p.name
            if name.startswith('.~') or p.suffix.lower() not in suffixes:
                continue
            if all(tok in name for tok in tokens):
                return str(p)
    return None


def _write_invoice_workbook(records: Sequence[InvoiceRecord], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = '明细表'
    ws.merge_cells('A1:L1')
    ws['A1'] = '报销单电子发票登记表（按月累计填写）'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    headers = ['序号', '发票类型（1)', '开票日期(2)', '发票代码(3)', '发票号码(4)',
               '货物或应税劳务、服务名称/项目名称(5)', '金额（6）', '税额（7）',
               '价税合计金额(8）', '报销/冲账', '状态', '来源文件']
    _write_headers(ws, 2, headers)
    for idx, rec in enumerate(records, start=1):
        row = idx + 3
        ws.cell(row, 1, idx)
        ws.cell(row, 2, rec.invoice_type)
        ws.cell(row, 3, rec.issue_date.replace('-', '.') if rec.issue_date else '')
        ws.cell(row, 4, rec.invoice_code)
        ws.cell(row, 5, rec.invoice_number)
        ws.cell(row, 6, rec.item_name)
        ws.cell(row, 7, rec.amount)
        ws.cell(row, 8, rec.tax)
        if rec.amount is not None and rec.tax is not None:
            ws.cell(row, 9, '=G%d+H%d' % (row, row))
        else:
            ws.cell(row, 9, rec.total)
        ws.cell(row, 10, '报销单')
        ws.cell(row, 11, _status_text(rec))
        ws.cell(row, 12, rec.source_file)
        for col in (7, 8, 9):
            ws.cell(row, col).number_format = '0.00'
        if rec.status in ('需复核', '待录入', '重复'):
            _shade_row(ws, row, 'FFF6E7' if rec.status != '重复' else 'FDECEA')
    _fit_columns(ws, {6: 30, 11: 28, 12: 55})
    wb.save(path)


def _write_trip_workbook(
    trips: Sequence[TripRecord],
    records: Sequence[InvoiceRecord],
    path: str,
    template_dir: str,
) -> None:
    template = _find_template(template_dir, ('外访', '行程'), ('.xlsx',))
    if template:
        wb = load_workbook(template)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '行程表'
        ws['A1'] = '行程表'
        ws['A2'] = '报销人：                            车牌号：'
        _write_headers(ws, 3, ['序号', '日期', '驾车人', '乘车人', '用车时长', '目的地',
                              '发车前公里数', '回至公司公里数', '公司用车公里数',
                              '加油费', '停车费', 'ETC通行费', '状态', '来源'])
    fees = _fees_by_date(records)
    start_row = 4
    for idx, trip in enumerate(trips, start=1):
        row = start_row + idx - 1
        day_fees = fees.get(trip.date, {})
        ws.cell(row, 1, idx)
        ws.cell(row, 2, trip.date)
        ws.cell(row, 6, trip.address)
        ws.cell(row, 7, trip.start_odometer)
        ws.cell(row, 8, trip.end_odometer)
        if trip.start_odometer is not None and trip.end_odometer is not None:
            ws.cell(row, 9, '=H%d-G%d' % (row, row))
        else:
            ws.cell(row, 9, trip.trip_km)
        ws.cell(row, 10, day_fees.get('油费', None))
        ws.cell(row, 11, day_fees.get('停车费', None))
        ws.cell(row, 12, day_fees.get('ETC通行费', None))
        ws.cell(row, 13, trip.status + (('：' + trip.error) if trip.error else ''))
        ws.cell(row, 14, '出发：%s\n结束：%s' % (trip.start_file, trip.end_file))
        if trip.status != '已确认':
            _shade_row(ws, row, 'FFF6E7')
    _fit_columns(ws, {2: 14, 6: 34, 13: 28, 14: 55})
    wb.save(path)


def _write_expense_workbook(records: Sequence[InvoiceRecord], path: str, template_dir: str) -> None:
    template = _find_template(template_dir, ('费用报销明细',), ('.xlsx',))
    if template:
        wb = load_workbook(template)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = '费用报销明细（付款）'
        ws['A1'] = '费用报销明细（付款）'
        _write_headers(ws, 2, ['序号', '申请日期', '月份', '申请地区', '申请人', '费用科目', '摘要', '金额'])

    groups: Dict[str, List[InvoiceRecord]] = defaultdict(list)
    for rec in records:
        if rec.total is not None:
            groups[rec.category or '其他'].append(rec)
    today = _dt.date.today().isoformat()
    row = 3
    for idx, category in enumerate(sorted(groups), start=1):
        rows = groups[category]
        total = round(sum((r.total or 0.0) for r in rows), 2)
        months = sorted({r.issue_date[:7] for r in rows if r.issue_date})
        ws.cell(row, 1, idx)
        ws.cell(row, 2, today)
        ws.cell(row, 3, months[0] if len(months) == 1 else '、'.join(months))
        ws.cell(row, 4, '')
        ws.cell(row, 5, '')
        ws.cell(row, 6, category)
        ws.cell(row, 7, '%s发票 %d 张' % (category, len(rows)))
        ws.cell(row, 8, total)
        ws.cell(row, 8).number_format = '0.00'
        row += 1
    _fit_columns(ws, {2: 14, 3: 16, 6: 16, 7: 28, 8: 12})
    wb.save(path)


def _fees_by_date(records: Sequence[InvoiceRecord]) -> Dict[str, Dict[str, float]]:
    fees: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for rec in records:
        if not rec.issue_date or rec.total is None:
            continue
        fees[rec.issue_date][rec.category or '其他'] += rec.total
    return {d: {k: round(v, 2) for k, v in cats.items()} for d, cats in fees.items()}


def _status_text(rec: InvoiceRecord) -> str:
    detail = rec.error or ''
    return rec.status + (('：' + detail) if detail else '')


def _write_headers(ws, row: int, headers: Sequence[str]) -> None:
    fill = PatternFill('solid', fgColor='F1F5F9')
    for idx, value in enumerate(headers, start=1):
        c = ws.cell(row, idx, value)
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _shade_row(ws, row: int, color: str) -> None:
    fill = PatternFill('solid', fgColor=color)
    for col in range(1, ws.max_column + 1):
        ws.cell(row, col).fill = fill


def _fit_columns(ws, special: Dict[int, int]) -> None:
    for col in range(1, ws.max_column + 1):
        width = special.get(col)
        if width is None:
            width = 12
            for row in range(1, min(ws.max_row, 20) + 1):
                val = ws.cell(row, col).value
                if val is not None:
                    width = max(width, min(24, len(str(val)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
