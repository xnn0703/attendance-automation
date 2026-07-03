# -*- coding: utf-8 -*-
"""engine.py 单元测试：覆盖 analyze、ot_val、配置解析等关键路径。
直接 python3 运行，无 pytest 依赖。"""
import sys, os, tempfile, shutil
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
import engine
import _make_sample

PASS = 0
FAIL = 0

def check(name, cond):
    global PASS, FAIL
    if cond:
        PASS += 1
        print('  ✓ %s' % name)
    else:
        FAIL += 1
        print('  ✗ %s' % name)


def test_ru30_rd30():
    print('\n[ru30 / rd30]')
    check('ru30(0) = 0', engine.ru30(0) == 0)
    check('ru30(1) = 30', engine.ru30(1) == 30)
    check('ru30(30) = 30', engine.ru30(30) == 30)
    check('ru30(31) = 60', engine.ru30(31) == 60)
    check('rd30(0) = 0', engine.rd30(0) == 0)
    check('rd30(29) = 0', engine.rd30(29) == 0)
    check('rd30(30) = 30', engine.rd30(30) == 30)
    check('rd30(59) = 30', engine.rd30(59) == 30)


def test_parse_day():
    print('\n[parse_day]')
    # 空打卡
    r = engine.parse_day('')
    check('空打卡 cnt=0', r['cnt'] == 0)
    check('空打卡 first=None', r['first'] is None)
    check('空打卡 le=None', r['le'] is None)
    check('空打卡 wq=False', r['wq'] is False)
    # 单次打卡
    r = engine.parse_day('09:00')
    check('单次 cnt=1', r['cnt'] == 1)
    check('单次 first=540', r['first'] == 540)
    check('单次 le=540', r['le'] == 540)
    # 双次打卡
    r = engine.parse_day('09:00\n18:30')
    check('双次 cnt=2', r['cnt'] == 2)
    check('双次 first=540', r['first'] == 540)
    check('双次 le=1110', r['le'] == 1110)
    # 外勤打卡
    r = engine.parse_day('09:00 外勤\n18:30 外勤')
    check('外勤 wq=True', r['wq'] is True)
    # 跨午夜打卡
    r = engine.parse_day('23:30\n01:00')
    check('跨午夜 le=1500', r['le'] == 1500)  # 60 + 1440


def test_ot_val():
    print('\n[ot_val]')
    # 工作日加班：18:30 后打卡
    p = {'cnt': 2, 'first': 540, 'le': 1200, 'wq': False}  # 09:00-20:00
    ot = engine.ot_val(p, True, False)
    check('工作日 09:00-20:00 OT=1.5h', abs(ot - 1.5) < 0.01)
    
    # 工作日加班 >=2h 减 0.5h
    p = {'cnt': 2, 'first': 540, 'le': 1290, 'wq': False}  # 09:00-21:30
    ot = engine.ot_val(p, True, False)
    check('工作日 09:00-21:30 OT=2.5h', abs(ot - 2.5) < 0.01)
    
    # 工作日加班 <2h 不减
    p = {'cnt': 2, 'first': 540, 'le': 1230, 'wq': False}  # 09:00-20:30
    ot = engine.ot_val(p, True, False)
    check('工作日 09:00-20:30 OT=1.5h', abs(ot - 1.5) < 0.01)
    
    # 节假日加班
    p = {'cnt': 2, 'first': 540, 'le': 1020, 'wq': False}  # 09:00-17:00
    ot = engine.ot_val(p, False, False)
    check('节假日 09:00-17:00 OT=6.5h', abs(ot - 6.5) < 0.01)  # 8h - 1.5h午休
    
    # 节假日加班 >8h 减 0.5h
    p = {'cnt': 2, 'first': 540, 'le': 1110, 'wq': False}  # 09:00-18:30
    ot = engine.ot_val(p, False, False)
    check('节假日 09:00-18:30 OT=8h', abs(ot - 8.0) < 0.01)  # 9.5h - 1.5h午休 = 8h, 不大于8h不减
    
    # 外勤工作日不计加班
    p = {'cnt': 2, 'first': 540, 'le': 1200, 'wq': True}
    ot = engine.ot_val(p, True, False)
    check('外勤工作日 OT=0', ot == 0.0)
    
    # 排除人员不计加班
    p = {'cnt': 2, 'first': 540, 'le': 1200, 'wq': False}
    ot = engine.ot_val(p, True, True)
    check('排除人员 OT=0', ot == 0.0)
    
    # 打卡 <2 次不计加班
    p = {'cnt': 1, 'first': 540, 'le': 540, 'wq': False}
    ot = engine.ot_val(p, True, False)
    check('单次打卡 OT=0', ot == 0.0)
    
    # 9:00 之前打卡，first 被 cap 到 540
    p = {'cnt': 2, 'first': 480, 'le': 1200, 'wq': False}  # 08:00-20:00
    ot = engine.ot_val(p, True, False)
    check('08:00-20:00 OT=1.5h (first capped at 540)', abs(ot - 1.5) < 0.01)


def test_day_style():
    print('\n[day_style]')
    cfg = engine.default_config()
    cls = {}
    win = {'s': 1, 'e': 31, 'j': None}
    
    # 正常出勤
    p = {'cnt': 2, 'first': 540, 'le': 1110, 'wq': False}
    fill, fred = engine.day_style(p, 'Y28099', 5, True, win, cfg, cls)
    check('正常出勤 fill=None', fill is None)
    check('正常出勤 fred=False', fred is False)
    
    # 迟到红字 (首卡 09:05)
    p = {'cnt': 2, 'first': 545, 'le': 1110, 'wq': False}
    fill, fred = engine.day_style(p, 'Y28099', 5, True, win, cfg, cls)
    check('迟到 09:05 fill=None', fill is None)
    check('迟到 09:05 fred=True', fred is True)
    
    # 迟到较重红底 (首卡 09:35)
    p = {'cnt': 2, 'first': 575, 'le': 1110, 'wq': False}
    fill, fred = engine.day_style(p, 'Y28099', 5, True, win, cfg, cls)
    check('迟到较重 09:35 fill=RED', fill == engine.RED)
    check('迟到较重 09:35 fred=False', fred is False)
    
    # 早退红底 (末卡 18:00)
    p = {'cnt': 2, 'first': 540, 'le': 1080, 'wq': False}
    fill, fred = engine.day_style(p, 'Y28099', 5, True, win, cfg, cls)
    check('早退 18:00 fill=RED', fill == engine.RED)
    
    # 缺卡绿底
    p = {'cnt': 1, 'first': 540, 'le': 540, 'wq': False}
    fill, fred = engine.day_style(p, 'Y28099', 5, True, win, cfg, cls)
    check('缺卡 fill=GREEN', fill == engine.GREEN)
    
    # 未出勤红底
    p = {'cnt': 0, 'first': None, 'le': None, 'wq': False}
    fill, fred = engine.day_style(p, 'Y28099', 5, True, win, cfg, cls)
    check('未出勤 fill=RED', fill == engine.RED)
    
    # 外勤蓝底
    p = {'cnt': 2, 'first': 540, 'le': 1110, 'wq': True}
    fill, fred = engine.day_style(p, 'Y28099', 5, True, win, cfg, cls)
    check('外勤 fill=BLUE', fill == engine.BLUE)
    
    # 休息日
    p = {'cnt': 0, 'first': None, 'le': None, 'wq': False}
    fill, fred = engine.day_style(p, 'Y28099', 5, False, win, cfg, cls)
    check('休息日 fill=None', fill is None)


def test_breaks_quan():
    print('\n[breaks_quan]')
    cls = {}
    
    # 正常出勤不破全勤
    p = {'cnt': 2, 'first': 540, 'le': 1110, 'wq': False}
    check('正常不破全勤', not engine.breaks_quan(p, 'Y28099', 5, True, cls))
    
    # 迟到破全勤
    p = {'cnt': 2, 'first': 545, 'le': 1110, 'wq': False}
    check('迟到破全勤', engine.breaks_quan(p, 'Y28099', 5, True, cls))
    
    # 早退破全勤
    p = {'cnt': 2, 'first': 540, 'le': 1080, 'wq': False}
    check('早退破全勤', engine.breaks_quan(p, 'Y28099', 5, True, cls))
    
    # 缺卡破全勤
    p = {'cnt': 1, 'first': 540, 'le': 540, 'wq': False}
    check('缺卡破全勤', engine.breaks_quan(p, 'Y28099', 5, True, cls))
    
    # 未出勤破全勤
    p = {'cnt': 0, 'first': None, 'le': None, 'wq': False}
    check('未出勤破全勤', engine.breaks_quan(p, 'Y28099', 5, True, cls))
    
    # 休息日不破全勤
    p = {'cnt': 0, 'first': None, 'le': None, 'wq': False}
    check('休息日不破全勤', not engine.breaks_quan(p, 'Y28099', 5, False, cls))


def test_config_parse():
    print('\n[配置文件解析]')
    d = tempfile.mkdtemp()
    try:
        # 空配置文件（只有注释）→ 应该返回空配置，不是默认值
        with open(os.path.join(d, 'kq_config.txt'), 'w') as f:
            f.write('# comment\n')
        cfg = engine.read_config(d)
        check('空配置 excl 为空', cfg['excl'] == set())
        check('空配置 strict 为空', cfg['strict'] == {})
        check('空配置 font_only 为空', cfg['font_only'] == set())
        
        # 无配置文件 → 返回默认值
        os.remove(os.path.join(d, 'kq_config.txt'))
        cfg = engine.read_config(d)
        check('无配置文件用默认值', cfg == engine.default_config())
        
        # 完整配置
        with open(os.path.join(d, 'kq_config.txt'), 'w') as f:
            f.write('OT_EXCLUDE=Y17074,Y28001\n')
            f.write('STRICT_LATE=Y28006:0905\n')
            f.write('LATE_FONT_ONLY=Y28001\n')
        cfg = engine.read_config(d)
        check('OT_EXCLUDE 解析', cfg['excl'] == {'Y17074', 'Y28001'})
        check('STRICT_LATE 解析', cfg['strict'] == {'Y28006': 545})
        check('LATE_FONT_ONLY 解析', cfg['font_only'] == {'Y28001'})
        
        # 格式错误的 STRICT_LATE
        with open(os.path.join(d, 'kq_config.txt'), 'w') as f:
            f.write('STRICT_LATE=Y28006:905\n')  # 缺少前导零
        cfg = engine.read_config(d)
        check('STRICT_LATE 补零 905→0905', cfg['strict'].get('Y28006') == 545)
        
        # classify 解析
        with open(os.path.join(d, 'kq_classify.txt'), 'w') as f:
            f.write('# comment\n')
            f.write('Y28001|9=B\n')
            f.write('Y28005|22=G\n')
        cls = engine.read_classify(d)
        check('classify 解析', cls == {'Y28001|9': 'B', 'Y28005|22': 'G'})
        
        # keep 解析
        with open(os.path.join(d, 'kq_keep.txt'), 'w') as f:
            f.write('Y28068, Y28099\n')
        keep = engine.read_keep(d)
        check('keep 解析', keep == {'Y28068', 'Y28099'})
    finally:
        shutil.rmtree(d)


def test_status_excel_fill():
    print('\n[status_excel_fill 自检]')
    # 与 day_style 一致性：late
    cell = {'status': 'late', 'late_font': True}
    fill, fred = engine.status_excel_fill(cell)
    check('late → fill=None, fred=True', fill is None and fred is True)
    
    # lateheavy
    cell = {'status': 'lateheavy', 'late_font': False}
    fill, fred = engine.status_excel_fill(cell)
    check('lateheavy → fill=RED', fill == engine.RED)
    
    # absent
    cell = {'status': 'absent', 'late_font': False}
    fill, fred = engine.status_excel_fill(cell)
    check('absent → fill=RED', fill == engine.RED)
    
    # miss
    cell = {'status': 'miss', 'late_font': False}
    fill, fred = engine.status_excel_fill(cell)
    check('miss → fill=GREEN', fill == engine.GREEN)
    
    # biz
    cell = {'status': 'biz', 'late_font': False}
    fill, fred = engine.status_excel_fill(cell)
    check('biz → fill=BLUE', fill == engine.BLUE)
    
    # field
    cell = {'status': 'field', 'late_font': False}
    fill, fred = engine.status_excel_fill(cell)
    check('field → fill=BLUE', fill == engine.BLUE)
    
    # pending (suggest=R)
    cell = {'status': 'pending', 'suggest': 'R'}
    fill, fred = engine.status_excel_fill(cell)
    check('pending R → fill=RED', fill == engine.RED)
    
    # pending (suggest=G)
    cell = {'status': 'pending', 'suggest': 'G'}
    fill, fred = engine.status_excel_fill(cell)
    check('pending G → fill=GREEN', fill == engine.GREEN)


def test_person_summary():
    print('\n[person_summary]')
    emp = {
        'ot': 2.5,
        'full': True,
        'cells': {
            1: {'status': 'normal'}, 2: {'status': 'late'}, 3: {'status': 'lateheavy'},
            4: {'status': 'early'}, 5: {'status': 'absent'}, 6: {'status': 'miss'},
            7: {'status': 'biz'}, 8: {'status': 'field'}, 9: {'status': 'pending'},
            10: {'status': 'rest'}, 11: {'status': 'pre'}, 12: {'status': 'post'},
        }
    }
    s = engine.person_summary(emp)
    check('ot=2.5', s['ot'] == 2.5)
    check('full=True', s['full'] is True)
    # work_days = normal+late+lateheavy+early+field+miss+biz = 7（absent 不算出勤）
    check('work_days=7', s['work_days'] == 7)
    check('late=1', s['counts']['late'] == 1)
    check('lateheavy=1', s['counts']['lateheavy'] == 1)
    check('early=1', s['counts']['early'] == 1)
    check('absent=1', s['counts']['absent'] == 1)
    check('miss=1', s['counts']['miss'] == 1)
    check('biz=1', s['counts']['biz'] == 1)
    check('field=1', s['counts']['field'] == 1)
    check('pending=1', s['counts']['pending'] == 1)


def test_out_name():
    print('\n[out_name]')
    check('全角原表 → 初表', engine.out_name('南京六部（原表）.xlsx', '（原表）', '（初表）') == '南京六部（初表）.xlsx')
    check('半角原表 → 初表', engine.out_name('南京六部(原表).xlsx', '（原表）', '（初表）') == '南京六部（初表）.xlsx')
    check('全角初表 → 考勤', engine.out_name('南京六部（初表）.xlsx', '（初表）', '（考勤）') == '南京六部（考勤）.xlsx')
    check('半角初表 → 考勤', engine.out_name('南京六部(初表).xlsx', '（初表）', '（考勤）') == '南京六部（考勤）.xlsx')


def test_shifted_tiao_headers():
    print('\n[调班表动态表头]')
    d = tempfile.mkdtemp(prefix='kq_tiao_')
    try:
        path = os.path.join(d, '调班表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '南京六部6月排班'

        base_work = {1, 2, 3, 4, 5, 8, 9, 10, 11, 12, 15, 16, 17, 18, 22, 23, 24, 25, 26, 29, 30}
        row, col = 3, 1
        for dd in range(1, 31):
            ws.cell(row, col, '%d%s' % (dd, '班' if dd in base_work else '休'))
            col += 1
            if col > 7:
                row += 1
                col = 1

        # 2026-06 真实文件里右侧明细列较旧模板右移：N 姓名、P 原日期、Q 现调日期。
        ws.cell(2, 13, '地区')
        ws.cell(2, 14, '姓名')
        ws.cell(2, 15, '工号')
        ws.cell(2, 16, '原日期/班次')
        ws.cell(2, 17, '现调日期/班次')
        ws.cell(3, 14, '张三')
        ws.cell(3, 15, 'Y01002')
        ws.cell(3, 16, '2026/6/6')
        ws.cell(4, 13, '南京')
        ws.cell(4, 14, '曹兆阳')
        ws.cell(4, 15, 'Y28002')
        ws.cell(4, 16, '2026/6/6\n2026/6/21\n2026/6/27\n2026/6/28')
        ws.cell(4, 17, '2026/6/1\n2026/6/2\n2026/6/3')
        wb.save(path)

        cal = engine.load_calendar(path, 2026, 6)
        check('右移表头能解析姓名与日期列', cal['swap'].get('曹兆阳') == {1, 2, 3, 6, 21, 27, 28})
        check('模板张三不进入调班名单', '张三' not in cal['swap'])
        check('6/1 基准班翻转为休', engine.is_work(cal, '曹兆阳', 1) is False)
        check('6/6 基准休翻转为班', engine.is_work(cal, '曹兆阳', 6) is True)
    finally:
        shutil.rmtree(d, ignore_errors=True)


def test_analyze_synthetic_data():
    print('\n[analyze 合成数据回归]')
    d = tempfile.mkdtemp(prefix='kq_engine_')
    try:
        _make_sample.make_yuan(os.path.join(d, '南京六部（原表）.xlsx'))
        _make_sample.make_roster(os.path.join(d, '员工花名册.xlsx'))
        _make_sample.make_tiao(os.path.join(d, '调班表.xlsx'))
        engine.prep(d, set())

        data = engine.analyze(d, {}, engine.default_config())
        check('analyze 返回 6 人', len(data['employees']) == 6)
        check('full_list 是姓名列表', data['stats']['full_list'] == ['程有梅'])
        check('full_set 是工号集合', data['stats']['full_set'] == {'Y28010'})
        check('pending=2', data['stats']['pending'] == 2)

        # analyze 的 status 反推 Excel 样式，必须与权威 day_style 逐格一致。
        ctx = engine._load_people(d)
        mismatches = []
        for p in ctx['ppl']:
            win = engine.window(ctx['roster'], p['bare'], ctx['y'], ctx['mo'], ctx['days'])
            emp = next(e for e in data['employees'] if e['gh'] == p['gh'])
            for dd in range(1, ctx['days'] + 1):
                if dd < win['s'] or dd > win['e']:
                    continue
                is_w = engine.is_work(ctx['cal'], p['bare'], dd)
                pi = engine.parse_day(p['dmap'][dd])
                gold = engine.day_style(pi, p['gh'], dd, is_w, win, engine.default_config(), {})
                mine = engine.status_excel_fill(emp['cells'][dd])
                if gold != mine:
                    mismatches.append((p['gh'], dd, gold, mine))
        check('status_excel_fill 与 day_style 逐格一致', mismatches == [])
    finally:
        shutil.rmtree(d, ignore_errors=True)


if __name__ == '__main__':
    test_ru30_rd30()
    test_parse_day()
    test_ot_val()
    test_day_style()
    test_breaks_quan()
    test_config_parse()
    test_status_excel_fill()
    test_person_summary()
    test_out_name()
    test_shifted_tiao_headers()
    test_analyze_synthetic_data()
    
    print('\n' + '=' * 50)
    print('结果：%d 通过，%d 失败' % (PASS, FAIL))
    if FAIL > 0:
        sys.exit(1)
    else:
        print('ALL TESTS PASSED')
