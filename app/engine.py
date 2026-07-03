# -*- coding: utf-8 -*-
"""南京六部考勤引擎 (Python + openpyxl)。规则照搬已验证的 kq_run.ps1，着色用 openpyxl Fill/Font。
对外三动作：prep / worklist / build。GUI 直接传 dict（保留名单 / 归类 / 配置），也可用 txt 读取助手。"""
import os, re, glob, copy, datetime, warnings
import openpyxl
from openpyxl.styles import PatternFill, Font

FONT_NAME = '新宋体'
RED, GREEN, BLUE, PURPLE = 'FFFF0000', 'FF92D050', 'FFDDEBF7', 'FFD09ECE'
DEPT_OK = '南京六部'
SUF = '（离职）'
LIZHI = '离职'
WAIQIN = '外勤'
# 6 列固定(姓名/考勤组/部门/工号/职位/UserId) + 当月天数
FIXED_COLS = 6
MAX_DAYS = 31
NCOL = FIXED_COLS + MAX_DAYS

# ---------- data dir (cross-platform default) ----------
def default_data_dir():
    """默认数据目录：优先环境变量 KQ_DIR；否则 Windows=D:\\考勤、其它平台=~/考勤。"""
    return os.environ.get('KQ_DIR') or (r'D:\考勤' if os.name == 'nt' else os.path.expanduser('~/考勤'))

# ---------- file discovery ----------
def find_inputs(d):
    def find(tok, *nots):
        for f in sorted(glob.glob(os.path.join(d, '*.xlsx'))):
            b = os.path.basename(f)
            if b.startswith('~$'):
                continue
            if tok in b and all(n not in b for n in nots):
                return f
        return None
    return {'yuan': find('原表', '初表', '考勤'), 'chu': find('初表'),
            'ros': find('花名册'), 'tiao': find('调班'), 'kao': find('考勤')}

# ---------- small helpers ----------
def s(v):
    return '' if v is None else str(v)

def to_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    try:
        n = float(v)
        return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=n)).date()
    except (ValueError, TypeError):
        return None

def in_month_day(dt, y, mo):
    if dt and dt.year == y and dt.month == mo:
        return dt.day
    return None

def year_month(ws):
    a1 = s(ws['A1'].value)
    m = re.search(r'(\d{4})-(\d{1,2})-\d{1,2}', a1)
    if not m:
        raise ValueError('A1 找不到 YYYY-MM 周期: ' + a1)
    return int(m.group(1)), int(m.group(2))

def month_days(y, mo):
    return (datetime.date(y + (mo // 12), (mo % 12) + 1, 1) - datetime.timedelta(days=1)).day

def month_ncol(y, mo):
    return FIXED_COLS + month_days(y, mo)

# ---------- roster ----------
def load_roster(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    r = {}
    for row in range(4, ws.max_row + 1):
        name = s(ws.cell(row, 3).value).strip()      # C 姓名
        if not name:
            continue
        e = {'name': name, 'zhiwei': s(ws.cell(row, 5).value), 'yong': s(ws.cell(row, 6).value),
             'gonghao': s(ws.cell(row, 7).value).strip(),
             'ruzhi': to_date(ws.cell(row, 8).value), 'lizhi': to_date(ws.cell(row, 9).value)}
        r.setdefault(name, []).append(e)
    wb.close()
    return r

def ros_of(roster, name):
    lst = roster.get(name)
    if not lst:
        return None
    act = [e for e in lst if e['yong'] != LIZHI]
    return act[0] if act else lst[0]

# ---------- calendar from 调班表 ----------
def load_calendar(path, y, mo):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    base_work, swap = {}, {}
    pat = re.compile(r'^\s*(\d{1,2})\s*([班休])')
    for row in range(1, ws.max_row + 1):
        for col in range(1, 8):                       # A-G = 周一..周日
            m = pat.match(s(ws.cell(row, col).value))
            if m and m.group(2) == '班':
                base_work[int(m.group(1))] = True
    def dates_in(txt):
        out = []
        dd = in_month_day(to_date(txt), y, mo)
        if dd:
            out.append(dd)
        # 兼容多种格式：2026/4/1、2026-04-01、datetime "2026-03-29 00:00:00"
        for mm in re.finditer(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})', s(txt)):
            if int(mm.group(1)) == y and int(mm.group(2)) == mo:
                d = int(mm.group(3))
                if d not in out:
                    out.append(d)
        return out

    def is_old_date_header(txt):
        txt = re.sub(r'\s+', '', s(txt))
        return '原' in txt and '日期' in txt

    def is_new_date_header(txt):
        txt = re.sub(r'\s+', '', s(txt))
        return ('现调' in txt or '现' in txt) and '日期' in txt

    def find_adjust_cols():
        # 右侧个人调班明细有过列位变化：旧模板 M/O/P，新模板 N/P/Q。
        # 按表头找列，找不到再回退到旧硬编码列，兼容历史文件。
        max_scan_row = min(ws.max_row, 10)
        for row in range(1, max_scan_row + 1):
            name_col = old_col = new_col = gh_col = None
            for col in range(8, ws.max_column + 1):
                txt = re.sub(r'\s+', '', s(ws.cell(row, col).value))
                if not txt:
                    continue
                if txt == '姓名' or (txt.endswith('姓名') and '系统姓名' not in txt):
                    name_col = col
                elif txt == '工号' or txt == '员工编号':
                    gh_col = col
                elif is_old_date_header(txt):
                    old_col = col
                elif is_new_date_header(txt):
                    new_col = col
            if name_col and old_col and new_col:
                return {'row': row, 'name': name_col, 'old': old_col, 'new': new_col, 'gh': gh_col}
        return {'row': 2, 'name': 13, 'old': 15, 'new': 16, 'gh': None}

    cols = find_adjust_cols()
    # 个人调班：原日期(O)+现调日期(P)两列涉及的当月日期，一律按基准日历"对调"班/休（翻转）。
    # 两列谁=班谁=休不固定，由基准日历决定：基准班→改休、基准休→改班。
    for row in range(cols['row'] + 1, ws.max_row + 1):
        nm = s(ws.cell(row, cols['name']).value).strip()
        if not nm:
            continue
        gh = s(ws.cell(row, cols['gh']).value).strip() if cols['gh'] else ''
        if nm == '张三' and gh == 'Y01002':
            continue
        for dd in dates_in(ws.cell(row, cols['old']).value) + dates_in(ws.cell(row, cols['new']).value):
            swap.setdefault(nm, set()).add(dd)
    wb.close()
    return {'base': base_work, 'swap': swap}

def is_work(cal, name, d):
    in_base = d in cal['base']
    if name in cal['swap'] and d in cal['swap'][name]:
        return not in_base                            # 调班日：翻转基准班/休
    return in_base

# ---------- punch parse ----------
def parse_day(txt):
    """解析某日打卡文本。
    
    返回:
        cnt: 打卡次数
        first: 首卡时间（分钟，0-1439）
        le: 末卡有效时间（分钟，跨午夜则 +1440）
        wq: 是否包含"外勤"标记
    """
    prev_minute = None
    midnight_offset = 0        # 跨午夜累加的 1440 倍数
    first_punch = None
    last_effective = None
    punch_count = 0
    is_waiqin = False
    for ln in s(txt).split('\n'):
        m = re.search(r'(\d{1,2}):(\d{2})', ln)
        if m:
            mm = int(m.group(1)) * 60 + int(m.group(2))
            if prev_minute is not None and mm < prev_minute:
                midnight_offset += 1440
            eff = mm + midnight_offset
            if punch_count == 0:
                first_punch = mm
            last_effective = eff; prev_minute = mm; punch_count += 1
        if WAIQIN in ln:
            is_waiqin = True
    return {'cnt': punch_count, 'first': first_punch, 'le': last_effective, 'wq': is_waiqin}

def ru30(m):
    """向上取整到 30 分钟。"""
    return ((m + 29) // 30) * 30

def rd30(m):
    """向下取整到 30 分钟。"""
    return (m // 30) * 30

def ot_val(p, is_w, is_excl):
    if is_excl or p['cnt'] < 2:
        return 0.0
    if is_w:
        if p['wq']:
            return 0.0
        end = rd30(p['le']); o = end - 1110
        if o < 0: o = 0
        if o >= 120: o -= 30
        return o / 60.0
    st = ru30(p['first'])
    if st < 540: st = 540
    er = p['le']
    if p['wq']: er = min(er, 1260)
    end = rd30(er); raw = end - st
    if raw < 0: raw = 0
    lunch = max(0, min(end, 810) - max(st, 720))
    o = raw - lunch
    if o > 480: o -= 30
    if o < 0: o = 0
    return o / 60.0

def fmt_ot(h):
    if h <= 0:
        return None
    r = round(h, 1)
    return str(int(r)) if r == int(r) else str(r)

# ---------- style decision -> (fill_color_or_None, font_red_bool) ----------
def day_style(p, gh, d, is_w, win, cfg, cls):
    if not is_w:
        if p['cnt'] >= 2 and p['wq']:
            return (BLUE, False)
        return (None, False)
    if p['cnt'] < 2:
        k = cls.get('%s|%d' % (gh, d)) or ('R' if p['cnt'] == 0 else 'G')
        return ({'R': RED, 'G': GREEN, 'B': BLUE}.get(k, BLUE), False)
    thr = cfg['strict'].get(gh, 570)
    if gh in cfg['font_only'] or (win['j'] is not None and d == win['j']):
        thr = 100000
    f = p['first']; le = p['le']
    if p['wq']:
        if f is not None and 540 < f <= thr:
            return (BLUE, True)
        return (BLUE, False)
    heavy = f is not None and f > thr
    mild = f is not None and f > 540 and not heavy
    early = le is not None and le < 1110
    if heavy or early:
        return (RED, True) if mild else (RED, False)
    if mild:
        return (None, True)
    return (None, False)

def breaks_quan(p, gh, d, is_w, cls):
    if not is_w:
        return False
    if p['cnt'] < 2:
        k = cls.get('%s|%d' % (gh, d)) or ('R' if p['cnt'] == 0 else 'G')
        return k in ('R', 'G')
    if p['first'] is not None and p['first'] > 540:
        return True
    if (not p['wq']) and p['le'] is not None and p['le'] < 1110:
        return True
    return False

def window(roster, name, y, mo, days):
    ros = ros_of(roster, name)
    win = {'s': 1, 'e': days, 'j': None}
    if ros:
        rz = in_month_day(ros['ruzhi'], y, mo)
        if rz:
            win['s'] = rz; win['j'] = rz
        lz = in_month_day(ros['lizhi'], y, mo)
        if lz:
            win['e'] = lz
    return win

# ---------- analyze 辅助 ----------
# status 取值（与设计稿语义对齐；底色/红字保证与 day_style 一致，见 _status_fill/自检）：
#   normal 正常 · late 迟到(红字) · lateheavy 迟到较重(红底) · early 早退(红底)
#   absent 未出勤(红底) · miss 缺卡(绿底) · biz 公出(蓝底) · field 外勤日(蓝底)
#   pending 待归类 · rest 休息/非工作日 · pre 入职前 · post 离职后
ANOMALY_STATUS = ('late', 'lateheavy', 'early', 'absent', 'miss', 'pending')
# status -> Excel 底色（None=不填），用于与 day_style 自检；pending 按建议值另算
_STATUS_FILL = {'absent': RED, 'miss': GREEN, 'biz': BLUE, 'lateheavy': RED, 'early': RED,
                'field': BLUE, 'late': None, 'normal': None, 'rest': None, 'pre': None, 'post': None}


def _hhmm(m):
    """分钟数转 HH:MM 字符串。"""
    if m is None:
        return None
    m %= 1440
    return '%02d:%02d' % (m // 60, m % 60)


def _punch_times(txt):
    """解析打卡文本，返回时间字符串列表。"""
    out = []
    for ln in s(txt).split('\n'):
        mt = re.search(r'(\d{1,2}):(\d{2})', ln)
        if mt:
            out.append('%02d:%02d' % (int(mt.group(1)), int(mt.group(2))))
    return out


def _late_threshold(gh, win, dd, cfg):
    """该格的迟到底色阈值（分钟）；font_only / 入职当天 → 极大值（仅红字、不上底色）。"""
    thr = cfg['strict'].get(gh, 570)
    if gh in cfg['font_only'] or (win['j'] is not None and dd == win['j']):
        thr = 100000
    return thr


# ---------- config / classify / keep (defaults = current Nanjing-Liu-Bu) ----------
def default_config():
    return {'excl': {'Y17074', 'Y28001'}, 'strict': {'Y28006': 545}, 'font_only': {'Y28001'}}

def read_config(d):
    cfg = default_config()
    f = os.path.join(d, 'kq_config.txt')
    if not os.path.exists(f):
        return cfg
    cfg = {'excl': set(), 'strict': {}, 'font_only': set()}
    with open(f, encoding='utf-8') as fh:
        for ln in fh:
            ln = ln.strip()
            if not ln or ln.startswith('#') or '=' not in ln:
                continue
            k, v = ln.split('=', 1); k = k.strip(); v = v.strip()
            if k == 'OT_EXCLUDE':
                cfg['excl'] |= {x.strip() for x in v.split(',') if x.strip()}
            elif k == 'LATE_FONT_ONLY':
                cfg['font_only'] |= {x.strip() for x in v.split(',') if x.strip()}
            elif k == 'STRICT_LATE':
                for pair in v.split(','):
                    if ':' in pair:
                        try:
                            g, hm = pair.split(':')
                            hm = hm.strip().zfill(4)  # 补零：905 → 0905
                            h, m_val = int(hm[:2]), int(hm[2:4])
                            if not (0 <= h <= 23 and 0 <= m_val <= 59):
                                raise ValueError(f'时间越界: {hm}')
                            cfg['strict'][g.strip()] = h * 60 + m_val
                        except (ValueError, IndexError) as e:
                            warnings.warn(f'跳过格式错误的 STRICT_LATE 条目 "{pair}": {e}')
    return cfg

def read_classify(d):
    c = {}
    f = os.path.join(d, 'kq_classify.txt')
    if os.path.exists(f):
        with open(f, encoding='utf-8') as fh:
            for ln in fh:
                ln = ln.strip()
                if ln and not ln.startswith('#') and '=' in ln:
                    k, v = ln.split('='); c[k.strip()] = v.strip().upper()
    return c

def read_keep(d):
    k = set()
    f = os.path.join(d, 'kq_keep.txt')
    if os.path.exists(f):
        with open(f, encoding='utf-8') as fh:
            for ln in fh:
                for g in re.split(r'[,\s]+', ln.strip()):
                    if g:
                        k.add(g)
    return k

# ---------- write config / classify / keep (与 skill 共用格式) ----------
def write_keep(d, keep):
    with open(os.path.join(d, 'kq_keep.txt'), 'w', encoding='utf-8') as f:
        f.write(','.join(sorted(keep)) + '\n')

def write_classify(d, classify):
    with open(os.path.join(d, 'kq_classify.txt'), 'w', encoding='utf-8') as f:
        f.write('# gonghao|day = B(公出) | G(缺卡) | R(未出勤)\n')
        for k in sorted(classify):
            f.write('%s=%s\n' % (k, classify[k]))

def write_config(d, config):
    with open(os.path.join(d, 'kq_config.txt'), 'w', encoding='utf-8') as f:
        f.write('OT_EXCLUDE=%s\n' % ','.join(sorted(config['excl'])))
        f.write('STRICT_LATE=%s\n' % ','.join('%s:%02d%02d' % (g, m // 60, m % 60) for g, m in config['strict'].items()))
        f.write('LATE_FONT_ONLY=%s\n' % ','.join(sorted(config['font_only'])))

# ---------- output filename ----------
def out_name(src_name, frm, to):
    for key in (frm, frm.replace('（', '(').replace('）', ')')):
        if key in src_name:
            return src_name.replace(key, to)
    base, ext = os.path.splitext(src_name)
    return base + to + ext


def validate(d):
    """返回输入问题清单（空列表=检查通过）。"""
    issues = []
    inp = find_inputs(d)
    if not inp['ros']:
        issues.append('缺少 *花名册*.xlsx')
    if not inp['tiao']:
        issues.append('缺少 *调班表*.xlsx')
    if not inp['yuan'] and not inp['chu']:
        issues.append('缺少 *原表*.xlsx（或已生成的 *初表*）')
    src = inp['chu'] or inp['yuan']
    if src:
        try:
            ws = openpyxl.load_workbook(src, data_only=True).active
            y, mo = year_month(ws)
            if inp['tiao']:
                tws = openpyxl.load_workbook(inp['tiao'], data_only=True).active
                tt = s(tws['A1'].value)
                m2 = re.search(r'(\d{1,2})\s*月', tt)
                if m2 and int(m2.group(1)) != mo:
                    issues.append('⚠ 月份不一致：原表/初表是 %d 月，调班表是 %d 月（"%s"）——请改用同一月份的原表与调班表！'
                                  % (mo, int(m2.group(1)), tt))
                cal = load_calendar(inp['tiao'], y, mo)
                if not cal['base']:
                    issues.append('调班表 A–G 列未解析出任何 N班/N休（请检查左侧排班网格）')
        except (ValueError, KeyError, openpyxl.utils.exceptions.InvalidFileException) as e:
            issues.append('解析失败：%s' % e)
    return issues

# ==================== STAGE: prep ====================
def prep(d, keep=None):
    keep = keep or set()
    inp = find_inputs(d)
    if not inp['yuan']:
        raise FileNotFoundError('缺少 *原表*.xlsx')
    if not inp['ros']:
        raise FileNotFoundError('缺少 *花名册*.xlsx')
    roster = load_roster(inp['ros'])
    wb = openpyxl.load_workbook(inp['yuan'])
    ws = wb.active
    y, mo = year_month(ws)
    ncol = month_ncol(y, mo)
    people = []; dropped = []; cand = []
    for r in range(5, ws.max_row + 1):
        name_raw = s(ws.cell(r, 1).value)
        if not name_raw.strip():
            continue
        dept = s(ws.cell(r, 3).value); name = name_raw.replace(SUF, '')
        pd = sum(1 for c in range(7, ncol + 1) if s(ws.cell(r, c).value).strip())
        if dept != DEPT_OK:
            dropped.append('%s [dept=%s]' % (name_raw, dept)); continue
        ros = ros_of(roster, name)
        if not ros:
            dropped.append('%s [无花名册匹配]' % name_raw); continue
        if ros['yong'] == LIZHI and pd < 7:
            cand.append({'name': name_raw, 'gh': ros['gonghao'], 'pd': pd, 'kept': ros['gonghao'] in keep})
            if ros['gonghao'] not in keep:
                dropped.append('%s [离职 pd=%d<7]' % (name_raw, pd)); continue
        cells = [(s(ws.cell(r, c).value) if ws.cell(r, c).value is not None else None,
                  copy.copy(ws.cell(r, c)._style)) for c in range(1, ncol + 1)]
        cells[3] = (ros['gonghao'], cells[3][1])      # D 工号
        cells[4] = (ros['zhiwei'], cells[4][1])       # E 职位
        people.append({'gh': ros['gonghao'], 'ghn': int(re.sub(r'\D', '', ros['gonghao']) or 0),
                       'name': name, 'yong': ros['yong'], 'pd': pd, 'cells': cells})
    people.sort(key=lambda p: p['ghn'])
    # rewrite person rows in place, sorted
    for i, p in enumerate(people):
        r = 5 + i
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c); val, st = p['cells'][c - 1]
            cell.value = val; cell._style = st
    extra = (ws.max_row - 4) - len(people)
    if extra > 0:
        ws.delete_rows(5 + len(people), extra)
    out = os.path.join(d, out_name(os.path.basename(inp['yuan']), '（原表）', '（初表）'))
    wb.save(out)
    return {'out': out, 'kept': people, 'dropped': dropped, 'lizhi_candidates': cand}

# ==================== shared load for worklist/build ====================
def _load_people(d):
    inp = find_inputs(d)
    if not inp['chu']:
        raise FileNotFoundError('缺少 *初表*.xlsx（先跑 prep）')
    if not inp['tiao']:
        raise FileNotFoundError('缺少 *调班表*.xlsx')
    if not inp['ros']:
        raise FileNotFoundError('缺少 *花名册*.xlsx')
    wb = openpyxl.load_workbook(inp['chu'])
    ws = wb.active
    y, mo = year_month(ws)
    days = month_days(y, mo)
    roster = load_roster(inp['ros'])
    cal = load_calendar(inp['tiao'], y, mo)
    ppl = []
    for r in range(5, ws.max_row + 1):
        name = s(ws.cell(r, 1).value)
        if not name.strip():
            continue
        gh = s(ws.cell(r, 4).value)
        dmap = {dd: s(ws.cell(r, dd + 6).value) for dd in range(1, days + 1)}
        ppl.append({'row': r, 'gh': gh, 'ghn': int(re.sub(r'\D', '', gh) or 0),
                    'name': name, 'bare': name.replace(SUF, ''), 'dmap': dmap})
    ppl.sort(key=lambda p: p['ghn'])
    return {'inp': inp, 'wb': wb, 'ws': ws, 'y': y, 'mo': mo, 'days': days,
            'roster': roster, 'cal': cal, 'ppl': ppl}

# ==================== STAGE: worklist ====================
def worklist(d):
    ctx = _load_people(d)
    cases = []
    for p in ctx['ppl']:
        win = window(ctx['roster'], p['bare'], ctx['y'], ctx['mo'], ctx['days'])
        for dd in range(1, ctx['days'] + 1):
            if dd < win['s'] or dd > win['e'] or not is_work(ctx['cal'], p['bare'], dd):
                continue
            txt = p['dmap'][dd]; pi = parse_day(txt)
            if pi['cnt'] < 2:
                cases.append({'gh': p['gh'], 'name': p['bare'], 'day': dd, 'cnt': pi['cnt'],
                              'wq': pi['wq'], 'punch': txt.replace('\n', ' | ').strip() or '(空)',
                              'suggest': 'R' if pi['cnt'] == 0 else 'G', 'key': '%s|%d' % (p['gh'], dd)})
    return {'year': ctx['y'], 'month': ctx['mo'], 'cases': cases}

# ==================== 公共计算层（build/analyze 共用） ====================
def _compute_plan(d, classify=None, config=None):
    """纯计算：逐人逐日判定 → 返回 plan/statis/pending，不触碰 Excel。
    
    返回:
        ctx: _load_people 的上下文
        cls/classify: 归类配置
        cfg/config: 策略配置
        plan: [{'p': ppl_entry, 'styles': {dd: (fill, fred)}, 'ot_by': {dd: ot}, 'quan': bool, 'cells': {dd: cell_info}}]
        summary: [{'gh', 'name', 'ot', 'quan', 'excl'}]
        pending: [{'gh', 'name', 'day', 'cnt', 'wq', 'punch', 'suggest', 'key'}]
        stats: {'people', 'counts', 'ot_total', 'pending', 'full_list', 'full_set', 'anomaly_set'}
    """
    ctx = _load_people(d)
    cls = classify if classify is not None else read_classify(d)
    cfg = config if config is not None else read_config(d)
    y, mo, days = ctx['y'], ctx['mo'], ctx['days']
    cal, roster = ctx['cal'], ctx['roster']
    
    plan = []
    summary = []
    pending = []
    counts = {k: 0 for k in ('late', 'lateheavy', 'early', 'absent', 'miss', 'biz', 'field')}
    ot_total = 0.0
    pending_total = 0
    full_set = set()
    anomaly_set = set()
    
    for p in ctx['ppl']:
        bare = p['bare']
        gh = p['gh']
        win = window(roster, bare, y, mo, days)
        ros = ros_of(roster, bare)
        is_excl = gh in cfg['excl']
        
        styles = {}
        ot_by = {}
        ot_sum = 0.0
        quan = True
        cells = {}
        has_attendance = False
        
        for dd in range(1, days + 1):
            txt = p['dmap'][dd]
            if dd < win['s'] or dd > win['e']:
                cells[dd] = {'status': 'pre' if dd < win['s'] else 'post'}
                continue
            
            is_w = is_work(cal, bare, dd)
            swap = (bare in cal['swap'] and dd in cal['swap'][bare])
            pi = parse_day(txt)
            punches = _punch_times(txt)
            ot = ot_val(pi, is_w, is_excl)
            
            cell = {'punches': punches, 'in': (punches[0] if punches else None),
                    'out': (punches[-1] if punches else None), 'ot': ot, 'wq': pi['wq'],
                    'swap': swap, 'late_font': False, 'cnt': pi['cnt']}
            
            # 样式判定（build 用）
            fill, fred = day_style(pi, gh, dd, is_w, win, cfg, cls)
            if fill or fred:
                styles[dd] = (fill, fred)
            
            # 加班统计
            if ot > 0:
                ot_by[dd] = ot
                ot_sum += ot
                ot_total += ot
            
            # 全勤判定
            if is_w and breaks_quan(pi, gh, dd, is_w, cls):
                quan = False
            
            # 状态判定（analyze 用）
            if not is_w:                                   # 非工作日
                if pi['cnt'] >= 2 and pi['wq']:
                    cell['status'] = 'field'
                    cell['reason'] = '节假日外勤打卡'
                elif pi['cnt'] >= 1:
                    cell['status'] = 'normal'
                    cell['reason'] = '休息日打卡（加班）' if ot > 0 else '休息日打卡'
                else:
                    cell['status'] = 'rest'
                    cell['reason'] = '休息'
            elif pi['cnt'] < 2:                            # 工作日 <2 次 → 归类/待定
                k = cls.get('%s|%d' % (gh, dd))
                sug = 'R' if pi['cnt'] == 0 else 'G'
                if k is None:
                    cell['status'] = 'pending'
                    cell['suggest'] = sug
                    cell['reason'] = '工作日打卡 %d 次，待归类（建议：%s）' % (
                        pi['cnt'], {'R': '未出勤', 'G': '缺卡', 'B': '公出'}[sug])
                    pending.append({'gh': gh, 'name': bare, 'day': dd, 'cnt': pi['cnt'],
                                    'wq': pi['wq'], 'punch': txt.replace('\n', ' | ').strip() or '(空)',
                                    'suggest': sug, 'key': '%s|%d' % (gh, dd)})
                else:
                    cell['status'] = {'R': 'absent', 'G': 'miss', 'B': 'biz'}.get(k, 'biz')
                    cell['reason'] = '工作日打卡 %d 次 → %s' % (
                        pi['cnt'], {'absent': '未出勤', 'miss': '缺卡', 'biz': '公出'}[cell['status']])
            else:                                          # 工作日 ≥2 次
                thr = _late_threshold(gh, win, dd, cfg)
                f, le = pi['first'], pi['le']
                mild = f is not None and 540 < f <= thr
                heavy = f is not None and f > thr
                early = le is not None and le < 1110
                cell['late_font'] = mild
                if pi['wq']:                               # 工作日外勤 → 蓝底（迟到仅红字，不判早退）
                    cell['status'] = 'field'
                    cell['reason'] = '外勤日' + ('（首卡 %s 迟到）' % _hhmm(f) if mild else '')
                elif heavy:
                    cell['status'] = 'lateheavy'
                    cell['reason'] = '首卡 %s 晚于 %s → 迟到较重（底色红）' % (_hhmm(f), _hhmm(thr))
                    if early:
                        cell['reason'] += '；末卡 %s 早于 18:30（早退）' % _hhmm(le)
                elif early:
                    cell['status'] = 'early'
                    cell['reason'] = '末卡 %s 早于 18:30 → 早退（底色红）' % _hhmm(le)
                    if mild:
                        cell['reason'] += '；首卡 %s 迟到（字体红）' % _hhmm(f)
                elif mild:
                    cell['status'] = 'late'
                    cell['reason'] = '首卡 %s 晚于 09:00 → 迟到（字体红）' % _hhmm(f)
                else:
                    cell['status'] = 'normal'
                    cell['reason'] = '正常出勤'
            
            st = cell['status']
            if st in ('normal', 'field') and pi['cnt'] >= 1:
                has_attendance = True
            if st in counts:
                counts[st] += 1
            if st == 'pending':
                pending_total += 1
            if st in ANOMALY_STATUS:
                anomaly_set.add(gh)
            
            cells[dd] = cell
        
        full = quan and has_attendance
        if full:
            full_set.add(gh)
        
        plan.append({
            'p': p, 'styles': styles, 'ot_by': ot_by, 'quan': quan, 'cells': cells,
            'win': win, 'ros': ros, 'is_excl': is_excl, 'has_attendance': has_attendance
        })
        summary.append({'gh': gh, 'name': bare, 'ot': ot_sum, 'quan': quan, 'excl': is_excl})
    
    summary.sort(key=lambda x: int(re.sub(r'\D', '', x['gh']) or 0))
    full_list = [e['p']['bare'] for e in plan if e['quan'] and e['has_attendance']]
    stats = {'people': len(plan), 'counts': counts, 'ot_total': ot_total,
             'pending': pending_total, 'full_list': full_list,
             'full_set': full_set, 'anomaly_set': anomaly_set}
    
    return {'ctx': ctx, 'cls': cls, 'cfg': cfg, 'plan': plan, 'summary': summary,
            'pending': pending, 'stats': stats}


# ==================== STAGE: build ====================
def build(d, classify=None, config=None):
    result = _compute_plan(d, classify, config)
    ctx = result['ctx']
    ws = ctx['ws']; days = ctx['days']
    ncol = FIXED_COLS + days
    base_font = Font(name=FONT_NAME, size=12, color=RED)
    cells_meta = []
    
    # 收集 cells_meta（工作日 <2 次的格子）
    for item in result['plan']:
        p = item['p']
        for dd in range(1, days + 1):
            if dd < item['win']['s'] or dd > item['win']['e']:
                continue
            is_w = is_work(ctx['cal'], item['p']['bare'], dd)
            if is_w and item['cells'].get(dd, {}).get('cnt', 0) < 2:
                cells_meta.append({'key': '%s|%d' % (p['gh'], dd), 'gh': p['gh'], 'name': p['bare'],
                                   'day': dd, 'row': 2 * p['row'] - 5, 'col': dd + 6})
    
    # 自底向上写 Excel（插入 OT 行时不偏移未处理行）
    for item in sorted(result['plan'], key=lambda x: x['p']['row'], reverse=True):
        p = item['p']; r = p['row']
        # 着色日格 + 姓名格
        for dd, (fill, fred) in item['styles'].items():
            cell = ws.cell(r, dd + 6)
            if fill:
                cell.fill = PatternFill('solid', fgColor=fill)
            if fred:
                cell.font = Font(name=FONT_NAME, size=12, color=RED)
        if item['quan']:
            ws.cell(r, 1).fill = PatternFill('solid', fgColor=PURPLE)
        # 插入 OT 行
        ws.insert_rows(r + 1)
        tmpl = copy.copy(ws.cell(r, 6)._style)        # F (UserId) = neutral base style
        for c in range(1, ncol + 1):
            oc = ws.cell(r + 1, c); oc._style = copy.copy(tmpl)
            if c == 1:
                oc.value = '加班时长'
            elif c >= 7 and (c - 6) in item['ot_by']:
                rv = round(item['ot_by'][c - 6], 1)
                oc.value = int(rv) if rv == int(rv) else rv
        if r in ws.row_dimensions and ws.row_dimensions[r].height:
            ws.row_dimensions[r + 1].height = ws.row_dimensions[r].height
    
    out = os.path.join(d, out_name(os.path.basename(ctx['inp']['chu']), '（初表）', '（考勤）'))
    ctx['wb'].save(out)
    return {'out': out, 'year': ctx['y'], 'month': ctx['mo'], 'summary': result['summary'],
            'cells': cells_meta, 'total_ot': result['stats']['ot_total'],
            'quan_count': len(result['stats']['full_list'])}


# ==================== STAGE: analyze (结构化导出，供 GUI 渲染) ====================
def analyze(d, classify=None, config=None):
    """读《初表》+花名册+调班表 → 结构化逐格判定（复用 day_style/ot_val 等），供 GUI 渲染。"""
    result = _compute_plan(d, classify, config)
    ctx = result['ctx']
    y, mo, days = ctx['y'], ctx['mo'], ctx['days']
    cal = ctx['cal']
    roster = ctx['roster']
    
    calendar = [{'day': dd, 'weekday': datetime.date(y, mo, dd).weekday(),
                 'is_work': dd in cal['base']} for dd in range(1, days + 1)]
    
    employees = []
    for item in result['plan']:
        p = item['p']
        ros = item['ros']
        employees.append({
            'gh': p['gh'], 'ghn': p['ghn'], 'name': p['bare'], 'row': p['row'],
            'pos': (ros['zhiwei'] if ros else ''), 'yong': (ros['yong'] if ros else ''),
            'join': in_month_day(ros['ruzhi'], y, mo) if ros else None,
            'leave': in_month_day(ros['lizhi'], y, mo) if ros else None,
            'win': item['win'], 'punch_days': sum(1 for cell in item['cells'].values()
                                                  if cell.get('cnt', 0) > 0),
            'leaver': bool(ros and ros['yong'] == LIZHI), 'is_excl': item['is_excl'],
            'ot': sum(item['ot_by'].values()), 'full': item['quan'] and item['has_attendance'],
            'cells': item['cells']
        })
    
    employees.sort(key=lambda e: e['ghn'])
    return {'year': y, 'month': mo, 'days': days, 'calendar': calendar,
            'employees': employees, 'pending': result['pending'], 'stats': result['stats']}


def status_excel_fill(cell):
    """把 analyze 的格 cell 反推 Excel (fill, font_red)，用于与 day_style 自检比对。
    pending 用建议值映射（与 build 未归类时一致）。"""
    st = cell.get('status')
    if st == 'pending':
        return ({'R': RED, 'G': GREEN}.get(cell.get('suggest'), GREEN), False)
    return (_STATUS_FILL.get(st), bool(cell.get('late_font')))


def person_summary(emp):
    """单人月度小结：出勤/加班/各类计数/全勤。"""
    c = {k: 0 for k in ('late', 'lateheavy', 'early', 'absent', 'miss', 'biz', 'field', 'pending')}
    work_days = 0
    for cell in emp['cells'].values():
        st = cell.get('status')
        if st in c:
            c[st] += 1
        if st in ('normal', 'late', 'lateheavy', 'early', 'field', 'miss', 'biz'):
            work_days += 1
    return {'ot': emp['ot'], 'counts': c, 'work_days': work_days, 'full': emp['full']}


if __name__ == '__main__':
    import sys
    dd = sys.argv[2] if len(sys.argv) > 2 else default_data_dir()
    stage = sys.argv[1] if len(sys.argv) > 1 else 'build'
    if stage == 'prep':
        rpt = prep(dd, read_keep(dd))
        print('[prep] out=%s kept=%d dropped=%d' % (rpt['out'], len(rpt['kept']), len(rpt['dropped'])))
        for c in rpt['lizhi_candidates']:
            print('  lizhi<7:', c['name'], c['gh'], 'pd=%d' % c['pd'], 'KEEP' if c['kept'] else 'DROP')
    elif stage == 'worklist':
        rpt = worklist(dd)
        print('[worklist] %d-%d cases=%d' % (rpt['year'], rpt['month'], len(rpt['cases'])))
        for c in rpt['cases']:
            print('  %s d%d cnt=%d %s sug=%s [%s]' % (c['gh'], c['day'], c['cnt'], 'WQ' if c['wq'] else '  ', c['suggest'], c['punch']))
    else:
        rpt = build(dd)
        print('[build] out=%s %d-%d total_ot=%g quan=%d' % (rpt['out'], rpt['year'], rpt['month'], rpt['total_ot'], rpt['quan_count']))
        for x in rpt['summary']:
            print('  %-8s %-7s OT=%5g quan=%d excl=%d' % (x['gh'], x['name'], x['ot'], x['quan'], x['excl']))
