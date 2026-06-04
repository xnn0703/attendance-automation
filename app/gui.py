# -*- coding: utf-8 -*-
"""南京六部考勤 GUI 上位机 (PySide6)。5 步向导，调用已验证的 engine。"""
import sys, os, datetime
from PySide6 import QtWidgets, QtCore, QtGui
import openpyxl
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import engine

WEEK = '一二三四五六日'
CLS_DISP = {'G': '缺卡', 'B': '公出', 'R': '未出勤'}
DISP_CLS = {v: k for k, v in CLS_DISP.items()}


def weekday_cn(y, m, d):
    return WEEK[datetime.date(y, m, d).weekday()]


def col_letter(n):
    s = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('南京六部考勤 上位机')
        self.resize(960, 680)
        self.dir = None
        self.keep = set()
        self.classify = {}
        self.cell_to_key = {}
        self.headless = False
        self.config = engine.default_config()
        self.worklist_data = None

        central = QtWidgets.QWidget(); self.setCentralWidget(central)
        v = QtWidgets.QVBoxLayout(central)
        topbar = QtWidgets.QHBoxLayout()
        bsave = QtWidgets.QPushButton('保存决策→目录'); bsave.clicked.connect(self.save_decisions)
        bload = QtWidgets.QPushButton('从目录载入决策'); bload.clicked.connect(self.load_decisions)
        topbar.addWidget(bsave); topbar.addWidget(bload)
        topbar.addWidget(QtWidgets.QLabel('（决策=保留名单+归类+配置，存为 kq_*.txt，与 skill 通用）')); topbar.addStretch(1)
        v.addLayout(topbar)
        self.tabs = QtWidgets.QTabWidget(); v.addWidget(self.tabs, 1)
        self.tabs.addTab(self._tab_dir(), '① 选目录')
        self.tabs.addTab(self._tab_prep(), '② 整理')
        self.tabs.addTab(self._tab_class(), '③ 归类')
        self.tabs.addTab(self._tab_config(), '④ 配置')
        self.tabs.addTab(self._tab_build(), '⑤ 生成')
        for i in (1, 2, 3, 4):
            self.tabs.setTabEnabled(i, False)

        self.logbox = QtWidgets.QPlainTextEdit(readOnly=True)
        self.logbox.setMaximumHeight(150)
        v.addWidget(QtWidgets.QLabel('日志：'))
        v.addWidget(self.logbox)

    def log(self, m):
        self.logbox.appendPlainText(str(m))
        QtWidgets.QApplication.processEvents()

    def err(self, e):
        self.log('错误：%s' % e)
        if not self.headless:
            QtWidgets.QMessageBox.critical(self, '错误', str(e))

    # ---------- ① select dir ----------
    def _tab_dir(self):
        w = QtWidgets.QWidget(); g = QtWidgets.QVBoxLayout(w)
        row = QtWidgets.QHBoxLayout()
        self.dir_edit = QtWidgets.QLineEdit(readOnly=True)
        b = QtWidgets.QPushButton('浏览…'); b.clicked.connect(self.browse)
        row.addWidget(QtWidgets.QLabel('数据目录：')); row.addWidget(self.dir_edit, 1); row.addWidget(b)
        g.addLayout(row)
        self.files_tbl = QtWidgets.QTableWidget(5, 2)
        self.files_tbl.setHorizontalHeaderLabels(['输入表', '识别结果'])
        self.files_tbl.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        g.addWidget(self.files_tbl)
        self.warn_lbl = QtWidgets.QLabel(''); self.warn_lbl.setWordWrap(True); self.warn_lbl.setStyleSheet('color:#b00;')
        g.addWidget(self.warn_lbl)
        self.btn_confirm = QtWidgets.QPushButton('确认目录，进入整理 →')
        self.btn_confirm.clicked.connect(self.confirm_dir); self.btn_confirm.setEnabled(False)
        g.addWidget(self.btn_confirm)
        return w

    def browse(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, '选择含三张输入表的目录', os.getcwd())
        if d:
            self.set_dir(d)

    def set_dir(self, d):
        self.dir = d; self.dir_edit.setText(d)
        inp = engine.find_inputs(d)
        labels = [('原表(钉钉打卡)', 'yuan'), ('员工花名册', 'ros'), ('调班表', 'tiao'),
                  ('初表(中间件)', 'chu'), ('考勤(成品)', 'kao')]
        for i, (lab, key) in enumerate(labels):
            self.files_tbl.setItem(i, 0, QtWidgets.QTableWidgetItem(lab))
            val = os.path.basename(inp[key]) if inp[key] else '✗ 未找到'
            self.files_tbl.setItem(i, 1, QtWidgets.QTableWidgetItem(('✓ ' + val) if inp[key] else val))
        ok = bool(inp['yuan'] and inp['ros'] and inp['tiao'])
        self.btn_confirm.setEnabled(ok)
        try:
            issues = engine.validate(d)
        except Exception as e:
            issues = ['校验异常：%s' % e]
        self.warn_lbl.setText(('⚠ ' + '；'.join(issues)) if issues else '✓ 输入检查通过')
        self.log('目录：%s（原表%s 花名册%s 调班表%s）%s' % (
            d, '✓' if inp['yuan'] else '✗', '✓' if inp['ros'] else '✗', '✓' if inp['tiao'] else '✗',
            ('｜⚠ ' + '；'.join(issues)) if issues else '｜✓ 检查通过'))
        return ok

    def confirm_dir(self):
        self.config = engine.read_config(self.dir)
        if os.path.exists(os.path.join(self.dir, 'kq_keep.txt')):
            self.keep = engine.read_keep(self.dir)
        if os.path.exists(os.path.join(self.dir, 'kq_classify.txt')):
            self.classify = engine.read_classify(self.dir)
        self.refresh_config_fields()
        if self.classify or self.keep:
            self.log('已载入既有决策：归类 %d、保留 %d' % (len(self.classify), len(self.keep)))
        self.tabs.setTabEnabled(1, True); self.tabs.setCurrentIndex(1)
        self.log('目录已确认，请运行整理。')

    # ---------- ② prep ----------
    def _tab_prep(self):
        w = QtWidgets.QWidget(); g = QtWidgets.QVBoxLayout(w)
        b = QtWidgets.QPushButton('运行整理（生成《初表》）'); b.clicked.connect(self.run_prep)
        g.addWidget(b)
        g.addWidget(QtWidgets.QLabel('纳入名单：'))
        self.kept_tbl = QtWidgets.QTableWidget(0, 4)
        self.kept_tbl.setHorizontalHeaderLabels(['工号', '姓名', '用工类型', '打卡天数'])
        g.addWidget(self.kept_tbl, 1)
        g.addWidget(QtWidgets.QLabel('离职且打卡<7天 —— 勾选要【保留】的（保留则只算到离职日；不勾=删除）：'))
        self.cand_tbl = QtWidgets.QTableWidget(0, 4)
        self.cand_tbl.setHorizontalHeaderLabels(['保留', '姓名', '工号', '打卡天数'])
        g.addWidget(self.cand_tbl)
        self.lbl_dropped = QtWidgets.QLabel('删除：—'); self.lbl_dropped.setWordWrap(True)
        g.addWidget(self.lbl_dropped)
        b2 = QtWidgets.QPushButton('应用保留并重生成《初表》，进入归类 →'); b2.clicked.connect(self.apply_keep)
        g.addWidget(b2)
        return w

    def _do_prep(self):
        rpt = engine.prep(self.dir, self.keep)
        self.kept_tbl.setRowCount(len(rpt['kept']))
        for i, p in enumerate(rpt['kept']):
            for j, val in enumerate([p['gh'], p['name'], p['yong'], str(p['pd'])]):
                self.kept_tbl.setItem(i, j, QtWidgets.QTableWidgetItem(val))
        self.cand_tbl.setRowCount(len(rpt['lizhi_candidates']))
        for i, c in enumerate(rpt['lizhi_candidates']):
            chk = QtWidgets.QTableWidgetItem(); chk.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            chk.setCheckState(QtCore.Qt.Checked if c['kept'] else QtCore.Qt.Unchecked)
            self.cand_tbl.setItem(i, 0, chk)
            for j, val in enumerate([c['name'], c['gh'], str(c['pd'])]):
                self.cand_tbl.setItem(i, j + 1, QtWidgets.QTableWidgetItem(val))
        self.lbl_dropped.setText('删除（%d）：%s' % (len(rpt['dropped']), '；'.join(rpt['dropped'])))
        self.log('整理完成：纳入 %d 人，删除 %d，离职<7天候选 %d。' % (
            len(rpt['kept']), len(rpt['dropped']), len(rpt['lizhi_candidates'])))
        return rpt

    def run_prep(self):
        try:
            self._do_prep()
        except Exception as e:
            self.err(e)

    def apply_keep(self):
        try:
            self.keep = set()
            for i in range(self.cand_tbl.rowCount()):
                if self.cand_tbl.item(i, 0).checkState() == QtCore.Qt.Checked:
                    self.keep.add(self.cand_tbl.item(i, 2).text())
            self._do_prep()
            self.log('保留名单：%s' % ('、'.join(self.keep) if self.keep else '（无，全部删除）'))
            self.tabs.setTabEnabled(2, True); self.tabs.setTabEnabled(3, True); self.tabs.setTabEnabled(4, True)
            self.tabs.setCurrentIndex(2)
        except Exception as e:
            self.err(e)

    # ---------- ③ classify ----------
    def _tab_class(self):
        w = QtWidgets.QWidget(); g = QtWidgets.QVBoxLayout(w)
        row = QtWidgets.QHBoxLayout()
        b = QtWidgets.QPushButton('列出待定项'); b.clicked.connect(self.run_worklist)
        b2 = QtWidgets.QPushButton('全部按建议'); b2.clicked.connect(self.apply_suggestions)
        row.addWidget(b); row.addWidget(b2); row.addStretch(1)
        g.addLayout(row)
        g.addWidget(QtWidgets.QLabel('工作日打卡<2次 —— 逐条定性（默认：0次=未出勤、1次=缺卡；连续空白/外勤单卡常为公出）：'))
        self.work_tbl = QtWidgets.QTableWidget(0, 6)
        self.work_tbl.setHorizontalHeaderLabels(['姓名', '日期', '周', '打卡次数', '打卡内容', '归类'])
        self.work_tbl.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.Stretch)
        g.addWidget(self.work_tbl, 1)
        b3 = QtWidgets.QPushButton('保存归类，进入生成 →'); b3.clicked.connect(self.save_classify)
        g.addWidget(b3)
        return w

    def run_worklist(self):
        try:
            wl = engine.worklist(self.dir); self.worklist_data = wl
            cases = wl['cases']
            self.work_tbl.setRowCount(len(cases))
            for i, c in enumerate(cases):
                wd = weekday_cn(wl['year'], wl['month'], c['day'])
                vals = [c['name'], '%d/%d' % (wl['month'], c['day']), wd, str(c['cnt']),
                        ('[外勤] ' if c['wq'] else '') + c['punch']]
                for j, val in enumerate(vals):
                    it = QtWidgets.QTableWidgetItem(val); it.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.work_tbl.setItem(i, j, it)
                combo = QtWidgets.QComboBox(); combo.addItems(['缺卡', '公出', '未出勤'])
                pre = self.classify.get(c['key'], c['suggest'])
                combo.setCurrentText(CLS_DISP.get(pre, CLS_DISP[c['suggest']]))
                combo.setProperty('key', c['key'])
                self.classify[c['key']] = pre
                combo.currentTextChanged.connect(lambda text, k=c['key']: self.classify.__setitem__(k, DISP_CLS[text]))
                self.work_tbl.setCellWidget(i, 5, combo)
            self.log('%d-%d 待定项 %d 条。' % (wl['year'], wl['month'], len(cases)))
        except Exception as e:
            self.err(e)

    def apply_suggestions(self):
        if not self.worklist_data:
            self.run_worklist(); return
        for i, c in enumerate(self.worklist_data['cases']):
            cb = self.work_tbl.cellWidget(i, 5)
            if cb:
                cb.setCurrentText(CLS_DISP[c['suggest']])
        self.log('已全部置为建议值。')

    def save_classify(self):
        if not self.worklist_data:
            self.err('请先「列出待定项」。'); return
        self.classify = {}
        for i in range(self.work_tbl.rowCount()):
            cb = self.work_tbl.cellWidget(i, 5)
            self.classify[cb.property('key')] = DISP_CLS[cb.currentText()]
        self.log('归类已保存（%d 条）。' % len(self.classify))
        self.tabs.setCurrentIndex(4)

    # ---------- ④ config ----------
    def _tab_config(self):
        w = QtWidgets.QWidget(); f = QtWidgets.QFormLayout(w)
        c = self.config
        self.cfg_excl = QtWidgets.QLineEdit(','.join(sorted(c['excl'])))
        self.cfg_strict = QtWidgets.QLineEdit(','.join('%s:%02d%02d' % (g, m // 60, m % 60) for g, m in c['strict'].items()))
        self.cfg_font = QtWidgets.QLineEdit(','.join(sorted(c['font_only'])))
        f.addRow('不计加班(工号,逗号)：', self.cfg_excl)
        f.addRow('严格迟到阈值(工号:HHMM)：', self.cfg_strict)
        f.addRow('迟到只红字(工号,逗号)：', self.cfg_font)
        b = QtWidgets.QPushButton('应用配置'); b.clicked.connect(self.apply_config)
        f.addRow(b)
        f.addRow(QtWidgets.QLabel('（默认即南京六部当前口径，通常无需改动。）'))
        return w

    def apply_config(self):
        try:
            excl = {x.strip() for x in self.cfg_excl.text().split(',') if x.strip()}
            font = {x.strip() for x in self.cfg_font.text().split(',') if x.strip()}
            strict = {}
            for pair in self.cfg_strict.text().split(','):
                pair = pair.strip()
                if ':' in pair:
                    gg, hm = pair.split(':'); strict[gg.strip()] = int(hm[:2]) * 60 + int(hm[2:4])
            self.config = {'excl': excl, 'strict': strict, 'font_only': font}
            self.log('配置已应用。')
        except Exception as e:
            self.err(e)

    def refresh_config_fields(self):
        c = self.config
        self.cfg_excl.setText(','.join(sorted(c['excl'])))
        self.cfg_strict.setText(','.join('%s:%02d%02d' % (g, m // 60, m % 60) for g, m in c['strict'].items()))
        self.cfg_font.setText(','.join(sorted(c['font_only'])))

    def save_decisions(self):
        if not self.dir:
            self.err('请先选目录'); return
        try:
            self.apply_config()
            engine.write_keep(self.dir, self.keep)
            engine.write_classify(self.dir, self.classify)
            engine.write_config(self.dir, self.config)
            self.log('已保存决策：kq_keep.txt / kq_classify.txt / kq_config.txt')
        except Exception as e:
            self.err(e)

    def load_decisions(self):
        if not self.dir:
            self.err('请先选目录'); return
        self.config = engine.read_config(self.dir)
        self.keep = engine.read_keep(self.dir)
        self.classify = engine.read_classify(self.dir)
        self.refresh_config_fields()
        self.log('已从目录载入决策：归类 %d、保留 %d（如需刷新表格请重跑整理/列待定项）' % (len(self.classify), len(self.keep)))

    # ---------- ⑤ build ----------
    def _tab_build(self):
        w = QtWidgets.QWidget(); g = QtWidgets.QVBoxLayout(w)
        b = QtWidgets.QPushButton('生成《考勤》'); b.clicked.connect(self.run_build)
        g.addWidget(b)
        self.sum_lbl = QtWidgets.QLabel('—'); g.addWidget(self.sum_lbl)
        split = QtWidgets.QSplitter(QtCore.Qt.Vertical)
        sw = QtWidgets.QWidget(); sl = QtWidgets.QVBoxLayout(sw); sl.setContentsMargins(0, 0, 0, 0)
        self.sum_tbl = QtWidgets.QTableWidget(0, 4)
        self.sum_tbl.setHorizontalHeaderLabels(['工号', '姓名', '加班(h)', '全勤'])
        sl.addWidget(QtWidgets.QLabel('汇总：')); sl.addWidget(self.sum_tbl)
        pw = QtWidgets.QWidget(); pl = QtWidgets.QVBoxLayout(pw); pl.setContentsMargins(0, 0, 0, 0)
        self.preview = QtWidgets.QTableWidget(0, 0)
        self.preview.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.preview.cellDoubleClicked.connect(self.on_preview_dblclick)
        pl.addWidget(QtWidgets.QLabel('彩色预览（颜色即最终着色；偶数行=加班时长）。双击 绿/蓝/红 的"缺卡格"可改判，即时重算：'))
        pl.addWidget(self.preview)
        split.addWidget(sw); split.addWidget(pw); split.setSizes([150, 450])
        g.addWidget(split, 1)
        row2 = QtWidgets.QHBoxLayout()
        brecalc = QtWidgets.QPushButton('重算（应用当前归类/配置）'); brecalc.clicked.connect(self.run_build)
        self.btn_open = QtWidgets.QPushButton('打开考勤表（Excel/WPS）'); self.btn_open.clicked.connect(self.open_result)
        self.btn_open.setEnabled(False)
        row2.addWidget(brecalc); row2.addWidget(self.btn_open)
        g.addLayout(row2)
        return w

    def _set_work_combo(self, key, disp):
        for i in range(self.work_tbl.rowCount()):
            cb = self.work_tbl.cellWidget(i, 5)
            if cb and cb.property('key') == key:
                cb.blockSignals(True); cb.setCurrentText(disp); cb.blockSignals(False)
                return

    def on_preview_dblclick(self, qr, qc):
        meta = self.cell_to_key.get((qr + 1, qc + 1))
        if not meta:
            return
        items = ['缺卡', '公出', '未出勤']
        cur = CLS_DISP.get(self.classify.get(meta['key'], ''), '')
        idx = items.index(cur) if cur in items else 0
        choice, ok = QtWidgets.QInputDialog.getItem(
            self, '改判', '%s  %d日  当前=%s\n选择新定性：' % (meta['name'], meta['day'], cur or '默认'),
            items, idx, False)
        if ok and choice:
            self.classify[meta['key']] = DISP_CLS[choice]
            self._set_work_combo(meta['key'], choice)
            self.log('改判 %s(%s) %d日 → %s，重算中…' % (meta['name'], meta['gh'], meta['day'], choice))
            self.run_build()

    def render_preview(self, path):
        ws = openpyxl.load_workbook(path).active
        maxr, maxc = ws.max_row, 37
        t = self.preview
        t.setRowCount(maxr); t.setColumnCount(maxc)
        t.setHorizontalHeaderLabels([col_letter(c) for c in range(1, maxc + 1)])
        t.setVerticalHeaderLabels([str(r) for r in range(1, maxr + 1)])
        for r in range(1, maxr + 1):
            for c in range(1, maxc + 1):
                cell = ws.cell(r, c)
                it = QtWidgets.QTableWidgetItem('' if cell.value is None else str(cell.value))
                try:
                    if cell.fill is not None and cell.fill.patternType == 'solid':
                        rgb = cell.fill.fgColor.rgb
                        if isinstance(rgb, str) and len(rgb) == 8 and rgb[2:].upper() not in ('FFFFFF', '000000'):
                            it.setBackground(QtGui.QColor(int(rgb[2:4], 16), int(rgb[4:6], 16), int(rgb[6:8], 16)))
                except Exception:
                    pass
                try:
                    fc = cell.font.color
                    if fc is not None and getattr(fc, 'rgb', None) == 'FFFF0000':
                        it.setForeground(QtGui.QColor(255, 0, 0))
                except Exception:
                    pass
                if (r, c) in self.cell_to_key:
                    it.setToolTip('双击改判：缺卡/公出/未出勤')
                t.setItem(r - 1, c - 1, it)
        t.resizeColumnsToContents()
        t.resizeRowsToContents()

    def run_build(self):
        try:
            QtWidgets.QApplication.setOverrideCursor(QtCore.Qt.WaitCursor)
            rpt = engine.build(self.dir, classify=self.classify or None, config=self.config)
            QtWidgets.QApplication.restoreOverrideCursor()
            self.out_path = rpt['out']
            self.sum_lbl.setText('%d-%d  共 %d 人，加班合计 %g h，全勤 %d 人。  →  %s' % (
                rpt['year'], rpt['month'], len(rpt['summary']), rpt['total_ot'],
                rpt['quan_count'], os.path.basename(rpt['out'])))
            self.sum_tbl.setRowCount(len(rpt['summary']))
            for i, x in enumerate(rpt['summary']):
                vals = [x['gh'], x['name'], ('—' if x['excl'] else '%g' % x['ot']), '✓' if x['quan'] else '']
                for j, val in enumerate(vals):
                    self.sum_tbl.setItem(i, j, QtWidgets.QTableWidgetItem(val))
            self.cell_to_key = {(c['row'], c['col']): c for c in rpt.get('cells', [])}
            self.render_preview(rpt['out'])
            self.btn_open.setEnabled(True)
            self.log('已生成：%s' % rpt['out'])
        except Exception as e:
            QtWidgets.QApplication.restoreOverrideCursor()
            self.err(e)

    def open_result(self):
        if getattr(self, 'out_path', None) and os.path.exists(self.out_path):
            os.startfile(self.out_path)


def _run_smoke(d):
    os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
    QtWidgets.QApplication([])
    w = MainWindow(); w.headless = True
    if not w.set_dir(d):
        print('SMOKE FAIL: 目录无效'); return 1
    w.confirm_dir(); w.run_prep(); w.apply_keep(); w.run_worklist(); w.save_classify(); w.run_build()
    ok = (w.preview.rowCount() == 54) and (len(w.cell_to_key) > 0) and os.path.exists(getattr(w, 'out_path', ''))
    print('SMOKE %s: preview=%d cells=%d out=%s' % (
        'OK' if ok else 'FAIL', w.preview.rowCount(), len(w.cell_to_key), getattr(w, 'out_path', '')))
    return 0 if ok else 1


def main():
    if '--smoke' in sys.argv:
        i = sys.argv.index('--smoke')
        d = sys.argv[i + 1] if i + 1 < len(sys.argv) else os.getcwd()
        sys.exit(_run_smoke(d))
    app = QtWidgets.QApplication(sys.argv)
    w = MainWindow(); w.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
