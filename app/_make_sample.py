# -*- coding: utf-8 -*-
"""生成一套合成的考勤输入三表（原表/花名册/调班表），覆盖各类考勤状态，
用于在没有真实隐私数据的机器上做端到端冒烟 / 演示。默认写到 KQ_DIR（缺省 ~/考勤_sample）。
用法：python3 app/_make_sample.py [目标目录]"""
import os, sys, datetime
import openpyxl

YEAR, MONTH = 2026, 5
DAYS = 31
# 基准班 19 天（2026-05，5/1=周五，劳动节连休，5/9 补班）
BASE_WORK = {6, 7, 8, 9, 11, 12, 13, 14, 15, 18, 19, 20, 21, 22, 25, 26, 27, 28, 29}

# 人员：(工号, 姓名, 职位, 用工类型, 入职, 离职, {日: 打卡文本})  —— 覆盖各状态
NORMAL = '08:50\n18:35'
PEOPLE = [
    ('Y17074', '沈楠', '专员', '转正', datetime.date(2023, 3, 1), None,
     {15: '09:45\n18:35'}),                                   # 迟到较重(>9:30) + 不计加班(excl)
    ('Y28001', '曹兆阳', '组长', '转正', datetime.date(2022, 1, 5), None,
     {7: '09:20\n18:35'}),                                    # 迟到(font_only→仅红字) + 不计加班
    ('Y28006', '刘莹', '专员', '转正', datetime.date(2023, 6, 1), None,
     {8: '09:10\n18:35'}),                                    # strict 09:05 → 迟到较重(底色红)
    ('Y28010', '程有梅', '专员', '转正', datetime.date(2022, 2, 1), None,
     {6: '08:50\n20:30'}),                                    # 全勤 + 加班 1.5h（紫名）
    ('Y28020', '钟俊', '专员', '转正', datetime.date(2023, 1, 1), None,
     {11: '09:00', 12: '08:50\n17:30', 13: '08:50外勤\n18:40外勤'}),  # 缺卡 / 早退 / 外勤
    ('Y28030', '高瑞', '专员', '转正', datetime.date(2023, 1, 1), None,
     {14: '', 30: '09:00\n18:40'}),                           # 未出勤(0次) / 5-30 个人调班来的班
]
SWAPS = {'高瑞': [30]}  # 个人调班：5/30 基准休 → 改班


def _ws(wb, title):
    ws = wb.active
    ws.title = title
    return ws


def make_yuan(path):
    wb = openpyxl.Workbook()
    ws = _ws(wb, '打卡时间')
    ws['A1'] = '南京六部考勤  2026-05-01 至 2026-05-31'
    ws['A2'] = '生成时间 2026-06-01'
    headers = ['姓名', '考勤组', '部门', '工号', '职位', 'UserId'] + [str(d) for d in range(1, DAYS + 1)]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
        ws.cell(4, c, h)
    r = 5
    for gh, name, pos, yong, ruzhi, lizhi, special in PEOPLE:
        ws.cell(r, 1, name)
        ws.cell(r, 2, '南京六部')
        ws.cell(r, 3, '南京六部')
        ws.cell(r, 6, 'uid_%s' % gh)
        for dd in range(1, DAYS + 1):
            if dd in special:
                val = special[dd]
            elif dd in BASE_WORK:
                val = NORMAL
            else:
                val = ''
            if val:
                ws.cell(r, dd + 6, val)
        r += 1
    wb.save(path)


def make_roster(path):
    wb = openpyxl.Workbook()
    ws = _ws(wb, 'Sheet1')
    ws['A1'] = '员工花名册'
    ws['A2'] = '填报单位：南京六部'
    for c, h in enumerate(['序号', '部门', '姓名', '系统姓名', '职位', '用工类型', '员工编号', '入职时间', '离职日期'], 1):
        ws.cell(3, c, h)
    r = 4
    for i, (gh, name, pos, yong, ruzhi, lizhi, _s) in enumerate(PEOPLE, 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, '南京六部')
        ws.cell(r, 3, name)
        ws.cell(r, 4, name)
        ws.cell(r, 5, pos)
        ws.cell(r, 6, yong)
        ws.cell(r, 7, gh)
        ws.cell(r, 8, ruzhi)
        if lizhi:
            ws.cell(r, 9, lizhi)
        r += 1
    wb.save(path)


def make_tiao(path):
    wb = openpyxl.Workbook()
    ws = _ws(wb, 'Sheet1')
    ws['A1'] = '2026 年 5 月 排班表'
    ws['A2'] = '星期一'
    # 左半区 A–G：把 1..31 顺序铺进 7 列网格，单元格 "D班 9:00-18:30" / "D休"
    row = 3
    col = 1
    for dd in range(1, DAYS + 1):
        txt = ('%d班 9:00-18:30' % dd) if dd in BASE_WORK else ('%d休' % dd)
        ws.cell(row, col, txt)
        col += 1
        if col > 7:
            col = 1
            row += 1
    # 右半区 L–Q：个人调班名单（M姓名=13, O原=15, P现=16）
    ws.cell(2, 13, '姓名')
    ws.cell(2, 15, '原日期')
    ws.cell(2, 16, '现调日期')
    rr = 3
    for name, days in SWAPS.items():
        ws.cell(rr, 13, name)
        ws.cell(rr, 16, '、'.join('2026/5/%d' % d for d in days))
        rr += 1
    wb.save(path)


def main():
    d = sys.argv[1] if len(sys.argv) > 1 else (os.environ.get('KQ_DIR') or os.path.expanduser('~/考勤_sample'))
    os.makedirs(d, exist_ok=True)
    make_yuan(os.path.join(d, '南京六部（原表）.xlsx'))
    make_roster(os.path.join(d, '员工花名册.xlsx'))
    make_tiao(os.path.join(d, '调班表.xlsx'))
    print('合成三表已写入：%s' % d)
    for f in sorted(os.listdir(d)):
        if f.endswith('.xlsx'):
            print('  -', f)
    return d


if __name__ == '__main__':
    main()
